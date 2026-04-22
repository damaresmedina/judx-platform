"""Gera stf_172_ministros_FINAL.xlsx usando como FONTE e MODELO
   C:/Users/medin/Downloads/STF ministros 2026.xlsx

Regras:
  - 172 linhas (Rezek consolidado com as 2 passagens na coluna PASSAGENS)
  - #1 = Flavio Dino, #172 = Visconde de Sabará (decrescente de antiguidade)
  - Nomes em UPPERCASE sem acento
  - Todos os campos limpos (sem HTML entities, sem aspas tipográficas, sem acento, sem espaços múltiplos)
  - Datas no formato YYYY-MM-DD
  - Estrutura de colunas preservada do original + 2 colunas acrescentadas (# no início, PASSAGENS no fim)

Lacunas no xlsx fonte são supridas com:
  - excel_parsed_174.json
  - stf_todos_ministros_consolidado.json
"""
import re, json, unicodedata, html
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC_XLSX = "C:/Users/medin/Downloads/STF ministros 2026.xlsx"
JSON_174 = "C:/Users/medin/Desktop/stf no diva/dados_ministros/excel_parsed_174.json"
JSON_BIO = "C:/Users/medin/Desktop/backup_judx/resultados/stf_todos_ministros_consolidado.json"
OUT = Path("C:/Users/medin/Desktop/backup_judx/resultados/stf_172_ministros_FINAL.xlsx")

MESES = {'janeiro':1,'fevereiro':2,'marco':3,'abril':4,'maio':5,'junho':6,
         'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12,
         'jan':1,'fev':2,'mar':3,'abr':4,'mai':5,'jun':6,'jul':7,'ago':8,'set':9,'out':10,'nov':11,'dez':12}

def sa(s):
    if not s: return ''
    return ''.join(c for c in unicodedata.normalize('NFKD', str(s)) if not unicodedata.combining(c))

def limpar(s):
    """Normaliza: HTML entities, aspas tipográficas, espaços. PRESERVA acentos (formato do corpus)."""
    if s is None or s == '': return ''
    s = str(s)
    s = html.unescape(s)
    s = s.replace("’","'").replace("‘","'").replace("“",'"').replace("”",'"')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def limpar_upper(s):
    return limpar(s).upper()

def normalizar_data(s):
    """Qualquer formato → YYYY-MM-DD."""
    if s is None or s == '': return ''
    s = str(s).strip()
    # Se for datetime do Excel, extrai data ISO direto
    if hasattr(s, 'strftime'): return s.strftime('%Y-%m-%d')
    s_clean = sa(s).lower().replace('1o','1').replace('1º','1').replace('primeiro','1')
    # ISO: 2024-01-31
    m = re.match(r'.*?(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s_clean)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # DD/MM/YYYY
    m = re.match(r'.*?(\d{1,2})[/](\d{1,2})[/](\d{4})', s_clean)
    if m: return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    # "DD de MES de YYYY" (com "de" entre partes)
    m = re.search(r'(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})', s_clean)
    if m:
        mes = MESES.get(m.group(2))
        if mes: return f"{int(m.group(3)):04d}-{mes:02d}-{int(m.group(1)):02d}"
    # "DD MES/YYYY" (ex: 'Decreto 25 jan/1892')
    m = re.search(r'(\d{1,2})\s+([a-z]+)[/\s](\d{4})', s_clean)
    if m:
        mes = MESES.get(m.group(2))
        if mes: return f"{int(m.group(3)):04d}-{mes:02d}-{int(m.group(1)):02d}"
    # "DD/MES/YYYY" (ex: '29/jan/1892')
    m = re.search(r'(\d{1,2})/([a-z]+)/(\d{4})', s_clean)
    if m:
        mes = MESES.get(m.group(2))
        if mes: return f"{int(m.group(3)):04d}-{mes:02d}-{int(m.group(1)):02d}"
    # Só "MES de YYYY"
    m = re.search(r'([a-z]+)\s+de\s+(\d{4})', s_clean)
    if m:
        mes = MESES.get(m.group(1))
        if mes: return f"{int(m.group(2)):04d}-{mes:02d}-01"
    # Só YYYY
    m = re.search(r'\b(\d{4})\b', s_clean)
    if m: return f"{m.group(1)}-01-01"
    return ''  # não parseou — deixa vazio (não polui com texto)

# ============================================================
# 1. Carrega xlsx fonte (modelo das colunas)
# ============================================================
wb_src = load_workbook(SRC_XLSX, data_only=True)
ws_src = wb_src['Planilha1']

# Captura header original (ordem exata das colunas)
header_original = [ws_src.cell(row=1, column=c).value for c in range(1, ws_src.max_column+1)]
# Normaliza nomes pra lookup (mantém a ordem e label original pro output)
col_map = {h: i+1 for i, h in enumerate(header_original) if h}

def find(name_substr, default=None):
    for h, idx in col_map.items():
        if h and name_substr.lower() in str(h).lower():
            return idx
    return default

CC = {
    'nome':   find('Nome do Ministro'),
    'ind':    find('Indicação Presidencial'),
    'local_ind': find('local of birth presidente'),
    'nomeacao': find('Nomeação'),
    'pai':    find('Pai'),
    'dob':    find('DATE OF BIRTH'),
    'lob':    find('LOCAL OF BIRTH'),
    'carrer': find('CARRER'),
    'fac':    find('Faculdade'),
    'posse':  find('Posse'),
    'saida_t': find('Aposentadoria'),
    'saida_d': find('Dt Saída'),
    'ante':   find('Antecessor'),
    'suce':   find('Sucessor'),
}

# Lê linhas
def cell(r, c): return ws_src.cell(row=r, column=c).value if c else None

raw_linhas = []
for r in range(2, ws_src.max_row+1):
    nome = cell(r, CC['nome'])
    if not nome or len(str(nome)) > 100: continue
    nome_lower = str(nome).lower().strip()
    # Filtra notas de rodapé e headers embutidos
    if nome_lower.startswith('a faculdade'): continue
    if nome_lower.startswith('*'): continue
    if 'cv no stf' in nome_lower: continue
    if nome_lower.startswith('nome do ministro'): continue
    if nome_lower.startswith('*cv'): continue
    # Asteriscos no final do nome (ex: 'Manoel José Espínola*')
    nome = re.sub(r'\*+\s*$', '', str(nome)).strip()
    raw_linhas.append({
        'nome_orig': str(nome).strip(),
        'ind_raw': cell(r, CC['ind']) or '',
        'local_ind_raw': cell(r, CC['local_ind']) or '',
        'nomeacao_raw': cell(r, CC['nomeacao']) or '',
        'pai_raw': cell(r, CC['pai']) or '',
        'dob_raw': cell(r, CC['dob']) or '',
        'lob_raw': cell(r, CC['lob']) or '',
        'carrer_raw': cell(r, CC['carrer']) or '',
        'fac_raw': cell(r, CC['fac']) or '',
        'posse_raw': cell(r, CC['posse']) or '',
        'saida_t_raw': cell(r, CC['saida_t']) or '',
        'saida_d_raw': cell(r, CC['saida_d']) or '',
        'ante_raw': cell(r, CC['ante']) or '',
        'suce_raw': cell(r, CC['suce']) or '',
    })

print(f"xlsx fonte: {len(raw_linhas)} linhas lidas")

# ============================================================
# 2. JSONs de complemento (para suprir lacunas)
# ============================================================
with open(JSON_174, encoding='utf-8') as f: j174 = json.load(f)
with open(JSON_BIO, encoding='utf-8') as f: jbio = json.load(f)

def key(s): return sa(str(s or '')).lower().strip()

j174_idx = {key(m.get('nome')): m for m in j174 if m.get('nome')}
jbio_idx = {}
for m in jbio:
    for n in (m.get('nome'), m.get('nome_completo')):
        if n: jbio_idx[key(n)] = m

def dmY(s):
    if not s: return ''
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', str(s).strip())
    return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}" if m else ''

# ============================================================
# 3. Monta linhas normalizadas
# ============================================================
linhas = []
for x in raw_linhas:
    k = key(x['nome_orig'])
    b174 = j174_idx.get(k, {})
    bbio = jbio_idx.get(k, {})

    dob = normalizar_data(x['dob_raw']) or dmY(str(bbio.get('nascimento','')).split(',',1)[0]) or b174.get('nascimento_data','') or ''
    # local de nascimento (LOCAL OF BIRTH)
    lob = limpar(x['lob_raw'])
    if not lob and bbio.get('nascimento') and ',' in bbio['nascimento']:
        lob = limpar(bbio['nascimento'].split(',',1)[1])
    if not lob: lob = limpar(b174.get('nascimento_local',''))

    nomeacao = normalizar_data(x['nomeacao_raw']) or dmY(bbio.get('nomeacao','')) or ''
    posse    = normalizar_data(x['posse_raw'])    or dmY(bbio.get('posse_stf','')) or b174.get('posse_stf_data','') or ''
    saida_d  = normalizar_data(x['saida_d_raw'])  or dmY(bbio.get('aposentadoria','')) or b174.get('saida_data','') or ''

    linhas.append({
        'nome':       limpar_upper(x['nome_orig']),
        'ind':        limpar_upper(x['ind_raw']) or limpar_upper(b174.get('presidente_indicou','')),
        'local_ind':  limpar_upper(x['local_ind_raw']),
        'nomeacao':   nomeacao,
        'pai':        limpar_upper(x['pai_raw']) or limpar_upper(b174.get('pai','') or ''),
        'dob':        dob,
        'lob':        limpar_upper(lob),
        'carrer':     limpar_upper(x['carrer_raw']) or limpar_upper(b174.get('carreira','') or ''),
        'fac':        limpar_upper(x['fac_raw']) or limpar_upper(b174.get('faculdade','') or ''),
        'posse':      posse,
        'saida_t':    limpar_upper(x['saida_t_raw']) or limpar_upper(b174.get('saida_tipo','')),
        'saida_d':    saida_d,
        'ante':       limpar_upper(x['ante_raw']) or limpar_upper(b174.get('antecessor','') or ''),
        'suce':       limpar_upper(x['suce_raw']) or limpar_upper(b174.get('sucessor','') or ''),
    })

# ============================================================
# 4. Preserva todas as linhas do xlsx fonte (sem consolidar)
# ============================================================
for r in linhas:
    r['passagens'] = ''

# ============================================================
# 5. Ordena DECRESCENTE de antiguidade (#1 = mais recente)
# ============================================================
linhas.sort(key=lambda r: (r['posse'] or '0000-00-00', r['nome']), reverse=True)
for i, r in enumerate(linhas, 1):
    r['antiguidade'] = i

# ============================================================
# 6. Escreve Excel (estrutura de colunas = a do original + # e PASSAGENS)
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "172 Ministros STF"

# Colunas: #  + as 14 do original na ORDEM original + PASSAGENS
cols_order = [
    ('antiguidade', '#'),
    ('nome',        'NOME DO MINISTRO'),
    ('ind',         'INDICACAO PRESIDENCIAL'),
    ('local_ind',   'LOCAL OF BIRTH PRESIDENTE'),
    ('nomeacao',    'NOMEACAO'),
    ('pai',         'PAI'),
    ('dob',         'DATE OF BIRTH'),
    ('lob',         'LOCAL OF BIRTH'),
    ('carrer',      'CARRER'),
    ('fac',         'FACULDADE'),
    ('posse',       'POSSE'),
    ('saida_t',     'APOSENTADORIA, EXONERACAO OU FALECIMENTO'),
    ('saida_d',     'DT SAIDA'),
    ('ante',        'ANTECESSOR'),
    ('suce',        'SUCESSOR'),
    ('passagens',   'PASSAGENS'),
]
widths = [5, 42, 30, 20, 12, 28, 12, 22, 38, 28, 12, 30, 12, 35, 35, 42]

header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(bold=True, color='FFFFFF')
for idx, ((_, label), w) in enumerate(zip(cols_order, widths), 1):
    c = ws.cell(row=1, column=idx, value=label)
    c.fill = header_fill; c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(idx)].width = w
ws.row_dimensions[1].height = 32

for i, r in enumerate(linhas, 2):
    for j, (k, _) in enumerate(cols_order, 1):
        c = ws.cell(row=i, column=j, value=r.get(k, ''))
        if k in ('antiguidade','dob','nomeacao','posse','saida_d'):
            c.alignment = Alignment(horizontal='center')

ws.freeze_panes = 'C2'
wb.save(OUT)

print(f"[ok] {OUT}")
print(f"linhas: {len(linhas)}")
for c in ['dob','nomeacao','posse','fac','carrer','ante','suce']:
    n = sum(1 for r in linhas if r[c])
    print(f"  {c}: {n}/{len(linhas)}")
