"""
backup-banco.py — Exporta TODAS as tabelas do JudX e ICONS para Excel
Gera um arquivo por banco com uma aba por tabela
Também exporta um resumo com contagens e último update

Usage: python scripts/backup-banco.py
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

JUDX_URL = 'postgresql://postgres:Zb9cHoRww7WxgT0C@db.ejwyguskoiraredinqmb.supabase.co:5432/postgres'
ICONS_URL = 'postgresql://postgres:RHuQvsf4shpsPRjP@db.hetuhkhhppxjliiaerlu.supabase.co:6543/postgres'

OUTPUT_DIR = r'C:\Users\medin\Desktop\backup_judx'
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADER_FONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1A2744', end_color='1A2744', fill_type='solid')
DATA_FONT = Font(name='Calibri', size=9)
THIN_BORDER = Border(
    bottom=Side(style='thin', color='DDDDDD')
)

def export_table(cursor, table_name, ws, limit=50000):
    """Export a table to a worksheet"""
    # Get columns
    cursor.execute(f"SELECT column_name FROM information_schema.columns WHERE table_schema='public' AND table_name='{table_name}' ORDER BY ordinal_position")
    columns = [r[0] for r in cursor.fetchall()]

    if not columns:
        ws.append(['Tabela vazia ou não encontrada'])
        return 0

    # Header
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Data
    cursor.execute(f"SELECT * FROM {table_name} LIMIT {limit}")
    rows = cursor.fetchall()

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            if isinstance(val, (dict, list)):
                val = str(val)[:500]  # Truncate JSON
            elif isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d %H:%M')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    # Auto-width (approximate)
    for ci, col in enumerate(columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max(len(col) + 4, 12), 50)

    # Freeze header
    ws.freeze_panes = 'A2'

    return len(rows)

def export_database(label, url, tables, filename):
    """Export entire database to Excel"""
    print(f'\n{"="*50}')
    print(f'  Exportando {label}...')
    print(f'{"="*50}')

    conn = psycopg2.connect(url, sslmode='require')
    cursor = conn.cursor()

    wb = openpyxl.Workbook()

    # Summary sheet first
    ws_summary = wb.active
    ws_summary.title = 'RESUMO'
    ws_summary.append(['Tabela', 'Registros', 'Exportado em'])
    for ci in range(1, 4):
        cell = ws_summary.cell(row=1, column=ci)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    summary_row = 2
    total_rows = 0

    for table_name in tables:
        print(f'  {table_name}...', end=' ', flush=True)

        # Count
        try:
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
            count = cursor.fetchone()[0]
        except Exception as e:
            print(f'ERRO: {e}')
            conn.rollback()
            continue

        # Create sheet (name max 31 chars)
        sheet_name = table_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Export
        exported = export_table(cursor, table_name, ws)
        total_rows += exported
        print(f'{exported} rows')

        # Summary
        ws_summary.cell(row=summary_row, column=1, value=table_name).font = DATA_FONT
        ws_summary.cell(row=summary_row, column=2, value=count).font = DATA_FONT
        ws_summary.cell(row=summary_row, column=3, value=datetime.now().strftime('%Y-%m-%d %H:%M')).font = DATA_FONT
        summary_row += 1

    # Total row
    ws_summary.cell(row=summary_row + 1, column=1, value='TOTAL').font = Font(bold=True)
    ws_summary.cell(row=summary_row + 1, column=2, value=f'=SUM(B2:B{summary_row-1})').font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 20

    output_path = os.path.join(OUTPUT_DIR, filename)
    wb.save(output_path)
    cursor.close()
    conn.close()

    print(f'\n  Salvo: {output_path} ({total_rows} rows total)')
    return output_path

def main():
    now = datetime.now().strftime('%Y%m%d_%H%M')

    # JudX tables
    judx_tables = [
        'stf_decisoes',
        'judx_case',
        'judx_decision',
        'stf_partes',
        'stj_temas',
        'stj_processos_semente',
        'stj_contramostra',
        'judx_court',
        'judx_judge',
        'judx_organ',
        'judx_procedural_class',
        'judx_subject',
    ]

    # ICONS tables
    icons_tables = [
        'objects',
        'edges',
    ]

    # Export JudX (stf_partes limited to 50K due to size)
    export_database('JudX', JUDX_URL, judx_tables, f'JudX_backup_{now}.xlsx')

    # Export ICONS (edges limited to 50K)
    export_database('ICONS', ICONS_URL, icons_tables, f'ICONS_backup_{now}.xlsx')

    print(f'\n{"="*50}')
    print(f'  Backups em: {OUTPUT_DIR}')
    print(f'{"="*50}')

if __name__ == '__main__':
    main()
