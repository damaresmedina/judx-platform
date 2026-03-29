"""
Auditoria STJ — conteúdo real: amostra de 20 registros por ano + estrutura + preenchimento
"""
import csv, os, glob, re, random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

dir_path = r"C:\Users\medin\Desktop\backup_judx\resultados\stj_datajud"
out_path = r"C:\Users\medin\Desktop\backup_judx\auditoria_stj_conteudo_v2_2026-03-29.xlsx"
files = sorted(glob.glob(os.path.join(dir_path, "*.csv")))

wb = Workbook()
hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(bold=True, color="FFFFFF", size=10)
alt_fill = PatternFill("solid", fgColor="D6E4F0")
yr_fill = PatternFill("solid", fgColor="FFC000")
thin = Side(style="thin")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_hdr(ws, headers, widths):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal="center"); c.border = brd
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

# ===================== ABA 1: AMOSTRA REAL =====================
ws = wb.active
ws.title = "Amostra Real"
display = ["ANO", "FONTE", "numero_processo", "classe", "data", "relator",
           "gabinete_ou_tipo", "assuntos", "ultima_fase_ou_teor", "movimentos", "formato", "grau"]
widths1 = [6, 8, 24, 20, 12, 28, 32, 45, 55, 10, 12, 8]
set_hdr(ws, display, widths1)

rn = 2
for fp in files:
    fname = os.path.basename(fp)
    m = re.search(r"(\d{4})", fname)
    ano = int(m.group(1)) if m else 0
    is_ckan = "ckan" in fname
    fonte = "CKAN" if is_ckan else "Datajud"

    rows_buf = []
    with open(fp, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader):
            rows_buf.append(row)
            if idx >= 2000:
                break

    if not rows_buf:
        continue

    n = min(20, len(rows_buf))
    sample = random.sample(rows_buf, n)

    # Header do ano
    for j in range(1, len(display) + 1):
        c = ws.cell(row=rn, column=j, value=f"--- {ano} ({fonte}) --- {len(rows_buf)}+ registros ---" if j == 1 else "")
        c.fill = yr_fill; c.font = Font(bold=True, size=9); c.border = brd
    ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=len(display))
    rn += 1

    for row in sample:
        if is_ckan:
            vals = [
                ano, fonte,
                row.get("numero_processo", ""),
                row.get("classe", ""),
                row.get("data_publicacao", ""),
                row.get("relator", ""),
                row.get("tipo_documento", ""),
                (row.get("assuntos", "") or "")[:80],
                (row.get("teor", "") or "")[:150],
                "", "", ""
            ]
        else:
            vals = [
                ano, fonte,
                row.get("numero_processo", ""),
                row.get("classe_nome", ""),
                row.get("data_ajuizamento", ""),
                row.get("relator", ""),
                (row.get("gabinete", "") or "")[:40],
                (row.get("assuntos", "") or "")[:80],
                (row.get("ultima_fase", "") or "")[:150],
                row.get("total_movimentos", ""),
                row.get("formato", ""),
                row.get("grau", "")
            ]

        for j, v in enumerate(vals):
            c = ws.cell(row=rn, column=j+1, value=v)
            c.border = brd
            c.font = Font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if rn % 2 == 0:
                c.fill = alt_fill
        rn += 1

print(f"Aba 1: {rn-1} linhas de amostra")

# ===================== ABA 2: ESTRUTURA =====================
ws2 = wb.create_sheet("Estrutura Campos")
ws2.cell(row=1, column=1, value="CAMPOS DATAJUD (2005-2025)").font = Font(bold=True, size=13)
ws2.cell(row=2, column=1, value="Fonte: api-publica.datajud.cnj.jus.br").font = Font(italic=True, size=10, color="666666")

datajud_fields = [
    ("numero_processo", "Numero CNJ do processo", "00045678920198100001"),
    ("classe_codigo", "Codigo numerico da classe", "11881 = AREsp, 1032 = REsp"),
    ("classe_nome", "Nome da classe processual", "Agravo em Recurso Especial"),
    ("data_ajuizamento", "Data de entrada no STJ", "2019-05-14"),
    ("relator", "Extraido do campo gabinete", "HERMAN BENJAMIN"),
    ("gabinete", "Gabinete original (fonte do relator)", "GABINETE DO MINISTRO HERMAN BENJAMIN"),
    ("orgao_julgador_codigo", "Codigo do orgao julgador", "1 = 1a Turma, 6 = Corte Especial"),
    ("assuntos", "Assuntos separados por ;", "Direito Tributario; ICMS; Creditamento"),
    ("ultima_fase", "Ultimo movimento processual", "Baixa Definitiva / Transito em Julgado"),
    ("total_movimentos", "Qtd movimentos no historico", "45"),
    ("formato", "Eletronico ou Fisico", "Eletronico"),
    ("grau", "Grau de jurisdicao", "SUP"),
    ("nivel_sigilo", "0=publico, >0=sigiloso", "0"),
    ("data_ultima_atualizacao", "Timestamp ultima movimentacao", "2024-03-15T14:30:00"),
]
for i, (campo, desc, ex) in enumerate(datajud_fields, 4):
    ws2.cell(row=i, column=1, value=campo).font = Font(bold=True, size=10)
    ws2.cell(row=i, column=1).border = brd
    ws2.cell(row=i, column=2, value=desc).font = Font(size=10)
    ws2.cell(row=i, column=2).border = brd
    ws2.cell(row=i, column=3, value=ex).font = Font(size=10, color="444444")
    ws2.cell(row=i, column=3).border = brd

r2 = len(datajud_fields) + 6
ws2.cell(row=r2, column=1, value="CAMPOS CKAN (2026)").font = Font(bold=True, size=13)
ws2.cell(row=r2+1, column=1, value="Fonte: dadosabertos.web.stj.jus.br").font = Font(italic=True, size=10, color="666666")

ckan_fields = [
    ("numero_processo", "Numero do processo", "00012345620268100001"),
    ("classe", "Sigla da classe", "AREsp"),
    ("data_publicacao", "Data publicacao no DJ", "15/03/2026"),
    ("data_recebimento", "Data de recebimento", "10/01/2026"),
    ("data_distribuicao", "Data distribuicao ao relator", "12/01/2026"),
    ("relator", "Ministro relator", "MINISTRO HERMAN BENJAMIN"),
    ("tipo_documento", "Tipo: acordao ou decisao mono", "DECISAO MONOCRATICA"),
    ("teor", "Texto da decisao (resumido)", "[texto livre, pode ter 500+ chars]"),
    ("descricao", "Descricao adicional", "[texto livre]"),
    ("assuntos", "Assuntos do processo", "Direito Civil; Contratos"),
    ("numero_registro", "Registro interno STJ", "2026/0012345-6"),
    ("recurso", "Tipo de recurso", "AREsp"),
]
for i, (campo, desc, ex) in enumerate(ckan_fields, r2+3):
    ws2.cell(row=i, column=1, value=campo).font = Font(bold=True, size=10)
    ws2.cell(row=i, column=1).border = brd
    ws2.cell(row=i, column=2, value=desc).font = Font(size=10)
    ws2.cell(row=i, column=2).border = brd
    ws2.cell(row=i, column=3, value=ex).font = Font(size=10, color="444444")
    ws2.cell(row=i, column=3).border = brd

ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 42
ws2.column_dimensions["C"].width = 42

# ===================== ABA 3: PREENCHIMENTO =====================
ws3 = wb.create_sheet("Preenchimento")
ws3.cell(row=1, column=1, value="Taxa de preenchimento por campo (amostra 1000 linhas por ano)").font = Font(bold=True, size=12)

ph = ["Ano", "N amostrado", "% processo", "% classe", "% data", "% relator", "% gabinete", "% assuntos", "% fase", "% movimentos", "% formato"]
for i, h in enumerate(ph, 1):
    c = ws3.cell(row=3, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd
ws3.column_dimensions["A"].width = 8
for i in range(2, 12):
    ws3.column_dimensions[get_column_letter(i)].width = 13

r3 = 4
fields_check = ["numero_processo", "classe_nome", "data_ajuizamento", "relator", "gabinete", "assuntos", "ultima_fase", "total_movimentos", "formato"]

for fp in files:
    fname = os.path.basename(fp)
    if "ckan" in fname:
        continue
    m = re.search(r"(\d{4})", fname)
    ano = int(m.group(1)) if m else 0

    counts = {k: 0 for k in fields_check}
    total = 0
    with open(fp, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if total >= 1000:
                break
            total += 1
            for k in counts:
                val = row.get(k, "").strip()
                if val and val != "0":
                    counts[k] += 1

    if total == 0:
        continue

    pcts = [round(v / total * 100, 1) for v in counts.values()]
    vals = [ano, total] + pcts
    for j, v in enumerate(vals):
        c = ws3.cell(row=r3, column=j+1, value=v)
        c.border = brd
        if r3 % 2 == 0:
            c.fill = alt_fill
        if j >= 2 and isinstance(v, float) and v < 50:
            c.font = Font(color="CC0000", bold=True)
        elif j >= 2 and isinstance(v, float) and v < 70:
            c.font = Font(color="FF8800")
    r3 += 1

# ===================== ABA 4: RESUMO =====================
ws4 = wb.create_sheet("Resumo")
grand = 0
for fp in files:
    count = 0
    with open(fp, "rb") as f:
        for _ in f:
            count += 1
    grand += count - 1

notas = [
    ("AUDITORIA DE CONTEUDO — CORPUS STJ", True, 14, None),
    ("29/03/2026 | Claude Code (JudX)", False, 10, "666666"),
    ("", False, 10, None),
    ("O QUE TEMOS:", True, 12, None),
    (f"  {grand:,} processos STJ em CSV local", False, 11, None),
    ("  Periodo: 2005-2026 (22 anos)", False, 11, None),
    ("  2 fontes: Datajud CNJ (2005-2025) + CKAN STJ (2026)", False, 11, None),
    ("", False, 10, None),
    ("ORGANIZACAO:", True, 12, None),
    ("  1 arquivo CSV por ano, nomeado stj_datajud_YYYY.csv ou stj_ckan_YYYY.csv", False, 11, None),
    ("  Pasta: Desktop/backup_judx/resultados/stj_datajud/", False, 11, None),
    ("  Datajud: 14 campos (processo, classe, data, relator, gabinete, assuntos, fase, movimentos...)", False, 11, None),
    ("  CKAN: 12 campos (processo, classe, datas, relator, tipo, teor da decisao, assuntos...)", False, 11, None),
    ("", False, 10, None),
    ("COBERTURA:", True, 12, None),
    ("  2005-2010: <1.000/ano — Datajud so tem retroativo parcial", False, 11, "CC0000"),
    ("  2011-2013: 3K-14K/ano — cobertura crescente", False, 11, "FF8800"),
    ("  2014-2016: 26K-50K/ano — cobertura robusta", False, 11, None),
    ("  2017-2024: 148K-362K/ano — cobertura massiva", False, 11, "006600"),
    ("  2025: 37K (pode estar incompleto)", False, 11, "FF8800"),
    ("  2026: 144K via CKAN (inclui teor da decisao)", False, 11, None),
    ("", False, 10, None),
    ("CAMPOS MAIS RICOS:", True, 12, None),
    ("  Datajud: relator preenchido em 60-94% (melhor 2012-2016)", False, 11, None),
    ("  Datajud: assuntos quase sempre presentes", False, 11, None),
    ("  Datajud: ultima_fase permite medir desfecho", False, 11, None),
    ("  CKAN 2026: tem TEOR da decisao (texto), unico ano com conteudo decisorio", False, 11, None),
    ("", False, 10, None),
    ("LIMITACOES:", True, 12, None),
    ("  Datajud e CKAN tem schemas diferentes — cruzamento exige normalizacao", False, 11, None),
    ("  Campo relator vazio em 30-75% dos registros recentes (2019+)", False, 11, None),
    ("  Nao ha inteiro teor no Datajud (so ultima_fase como proxy de resultado)", False, 11, None),
    ("  Anos pre-2011 sao amostra, nao universo", False, 11, None),
]

for i, (txt, bold, sz, color) in enumerate(notas):
    c = ws4.cell(row=i+1, column=1, value=txt)
    c.font = Font(bold=bold, size=sz, color=color or "000000")
ws4.column_dimensions["A"].width = 80

wb.save(out_path)
print(f"\nSalvo: {out_path}")
print(f"4 abas: Amostra Real, Estrutura Campos, Preenchimento, Resumo")
print(f"Total corpus: {grand:,} processos")
