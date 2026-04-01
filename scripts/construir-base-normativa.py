"""
Constroi base normativa: baixa codigos do planalto, extrai artigos e salva em CSV + Excel.
Cada artigo vira uma linha com: sigla, nome_codigo, lei_referencia, artigo, caput, dispositivos (incisos/paragrafos).
"""
import urllib.request, re, html as html_mod, csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

out_dir = r"C:\Users\medin\Desktop\backup_judx\resultados"
out_csv = os.path.join(out_dir, "base_normativa_codigos_2026-03-29.csv")
out_xlsx = os.path.join(out_dir, "base_normativa_codigos_2026-03-29.xlsx")

codigos = [
    ("CF88", "Constituicao Federal", "https://www.planalto.gov.br/ccivil_03/constituicao/constituicao.htm", "CF/1988"),
    ("CC", "Codigo Civil", "https://www.planalto.gov.br/ccivil_03/leis/2002/l10406compilada.htm", "Lei 10.406/2002"),
    ("CPC", "Codigo de Processo Civil", "https://www.planalto.gov.br/ccivil_03/_ato2015-2018/2015/lei/l13105.htm", "Lei 13.105/2015"),
    ("CPP", "Codigo de Processo Penal", "https://www.planalto.gov.br/ccivil_03/decreto-lei/del3689compilado.htm", "DL 3.689/1941"),
    ("CP", "Codigo Penal", "https://www.planalto.gov.br/ccivil_03/decreto-lei/del2848compilado.htm", "DL 2.848/1940"),
    ("CLT", "Consolidacao das Leis do Trabalho", "https://www.planalto.gov.br/ccivil_03/decreto-lei/del5452compilado.htm", "DL 5.452/1943"),
    ("CTN", "Codigo Tributario Nacional", "https://www.planalto.gov.br/ccivil_03/leis/l5172compilado.htm", "Lei 5.172/1966"),
    ("CDC", "Codigo de Defesa do Consumidor", "https://www.planalto.gov.br/ccivil_03/leis/l8078compilado.htm", "Lei 8.078/1990"),
    ("ECA", "Estatuto da Crianca e Adolescente", "https://www.planalto.gov.br/ccivil_03/leis/l8069.htm", "Lei 8.069/1990"),
    ("LEF", "Lei de Execucao Fiscal", "https://www.planalto.gov.br/ccivil_03/leis/l6830.htm", "Lei 6.830/1980"),
    ("LINDB", "Lei de Introducao ao Direito Brasileiro", "https://www.planalto.gov.br/ccivil_03/decreto-lei/del4657compilado.htm", "DL 4.657/1942"),
    ("LRF", "Lei de Responsabilidade Fiscal", "https://www.planalto.gov.br/ccivil_03/leis/lcp/lcp101.htm", "LC 101/2000"),
    ("LAI", "Lei de Acesso a Informacao", "https://www.planalto.gov.br/ccivil_03/_ato2011-2014/2011/lei/l12527.htm", "Lei 12.527/2011"),
    ("LIA", "Lei de Improbidade Administrativa", "https://www.planalto.gov.br/ccivil_03/leis/l8429.htm", "Lei 8.429/1992"),
    ("LEP", "Lei de Execucao Penal", "https://www.planalto.gov.br/ccivil_03/leis/l7210compilado.htm", "Lei 7.210/1984"),
    ("LOMAN", "Lei Organica da Magistratura", "https://www.planalto.gov.br/ccivil_03/leis/lcp/lcp35.htm", "LC 35/1979"),
    ("LOMP", "Lei Organica do Ministerio Publico", "https://www.planalto.gov.br/ccivil_03/leis/l8625.htm", "Lei 8.625/1993"),
]

all_arts = []

for sigla, nome, url, lei_ref in codigos:
    print(f"\nBaixando {sigla} ({nome})...")
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        resp = urllib.request.urlopen(req, timeout=30)
        raw = resp.read()
        # Tentar decodificar
        for enc in ["utf-8", "latin-1", "cp1252"]:
            try:
                html = raw.decode(enc)
                break
            except:
                continue
        else:
            html = raw.decode("utf-8", errors="replace")

        # Limpar
        text = re.sub(r"<[^>]+>", " ", html)
        text = html_mod.unescape(text)
        text = text.replace("\xa0", " ")
        text = re.sub(r"\s+", " ", text)

        # Extrair artigos
        # Pattern: "Art. N" ou "Art. N-A" etc
        art_splits = re.split(r"(Art\.\s*\d+[\-A-Z]*)", text)

        current_art = None
        count = 0
        for part in art_splits:
            m = re.match(r"Art\.\s*(\d+[\-A-Z]*)", part.strip())
            if m:
                current_art = m.group(1).replace("-", "").strip()
                continue
            if current_art is None:
                continue

            # Pegar o caput (primeira frase antes de inciso/paragrafo)
            caput = part.strip()[:500]
            # Limpar
            caput = re.sub(r"\s+", " ", caput).strip()
            if len(caput) < 5:
                continue

            # Contar dispositivos
            n_incisos = len(re.findall(r"\b[IVXLCDM]+\s*[-\u2013]", part))
            n_paragrafos = len(re.findall(r"[Pp]ar.grafo|\xa7|\xc2\xa7|Par.grafo .nico", part))
            n_alineas = len(re.findall(r"\b[a-z]\)", part))

            # Slug para ancoragem
            slug = f"{sigla.lower()}-art-{current_art}".lower()

            all_arts.append({
                "slug": slug,
                "sigla": sigla,
                "nome_codigo": nome,
                "lei_referencia": lei_ref,
                "artigo": current_art,
                "caput": caput[:300],
                "incisos": n_incisos,
                "paragrafos": n_paragrafos,
                "alineas": n_alineas,
            })
            count += 1

        print(f"  {count} artigos extraidos")

    except Exception as e:
        print(f"  ERRO: {e}")

# Deduplica por slug (pegar o primeiro de cada)
seen = set()
unique = []
for a in all_arts:
    if a["slug"] not in seen:
        seen.add(a["slug"])
        unique.append(a)

print(f"\nTotal: {len(all_arts)} artigos, {len(unique)} unicos")

# Salvar CSV
with open(out_csv, "w", encoding="utf-8", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["slug", "sigla", "nome_codigo", "lei_referencia", "artigo", "caput", "incisos", "paragrafos", "alineas"])
    w.writeheader()
    w.writerows(unique)
print(f"CSV: {out_csv}")

# Salvar Excel
wb = Workbook()
hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(bold=True, color="FFFFFF", size=10)
alt = PatternFill("solid", fgColor="D6E4F0")
thin = Side(style="thin")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

# ABA 1: Todos os artigos
ws = wb.active
ws.title = "Base Normativa"
hdrs = ["Slug", "Sigla", "Codigo", "Lei", "Artigo", "Caput (300 chars)", "Incisos", "Paragrafos", "Alineas"]
for i, h in enumerate(hdrs, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd
for i, w in enumerate([28, 8, 35, 18, 8, 80, 8, 10, 8]):
    ws.column_dimensions[get_column_letter(i+1)].width = w

for i, a in enumerate(unique):
    r = i + 2
    vals = [a["slug"], a["sigla"], a["nome_codigo"], a["lei_referencia"], a["artigo"], a["caput"], a["incisos"], a["paragrafos"], a["alineas"]]
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=j+1, value=v)
        c.border = brd
        c.alignment = Alignment(wrap_text=True, vertical="top") if j == 5 else Alignment(vertical="top")
        if i % 2 == 1: c.fill = alt

# ABA 2: Resumo por codigo
ws2 = wb.create_sheet("Resumo por Codigo")
for i, h in enumerate(["Sigla", "Codigo", "Lei", "Artigos", "Incisos total", "Paragrafos total"], 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd
ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 14

from collections import Counter
sigla_stats = {}
for a in unique:
    s = a["sigla"]
    if s not in sigla_stats:
        sigla_stats[s] = {"nome": a["nome_codigo"], "lei": a["lei_referencia"], "arts": 0, "inc": 0, "par": 0}
    sigla_stats[s]["arts"] += 1
    sigla_stats[s]["inc"] += a["incisos"]
    sigla_stats[s]["par"] += a["paragrafos"]

r2 = 2
total_arts = 0
for s in ["CF88", "CC", "CPC", "CPP", "CP", "CLT", "CTN", "CDC", "ECA", "LEF", "LINDB", "LRF", "LAI", "LIA", "LEP", "LOMAN", "LOMP"]:
    if s in sigla_stats:
        st = sigla_stats[s]
        total_arts += st["arts"]
        for j, v in enumerate([s, st["nome"], st["lei"], st["arts"], st["inc"], st["par"]]):
            c = ws2.cell(row=r2, column=j+1, value=v)
            c.border = brd
            if r2 % 2 == 0: c.fill = alt
        r2 += 1

# Total
tf = PatternFill("solid", fgColor="FFC000")
for j, v in enumerate(["TOTAL", "", "", total_arts, "", ""]):
    c = ws2.cell(row=r2, column=j+1, value=v)
    c.font = Font(bold=True); c.fill = tf; c.border = brd

wb.save(out_xlsx)
print(f"Excel: {out_xlsx}")
print(f"\n{len(unique)} artigos de {len(sigla_stats)} codigos/leis")
