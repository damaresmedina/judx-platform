"""
Contraste STJ: Justica em Numeros x Datajud API x CSVs locais
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

fp = r"C:\Users\medin\Desktop\backup_judx\dados-jn-cnj\JN_23-Set-2025.csv"
out = r"C:\Users\medin\Desktop\backup_judx\contraste_stj_cobertura_2026-03-29.xlsx"

# Ler STJ do Justica em Numeros
stj = {}
with open(fp, "r", encoding="utf-8", errors="replace") as f:
    reader = csv.DictReader(f, delimiter=";")
    for row in reader:
        if row.get("sigla") == "STJ":
            ano = int(row["ano"])
            def to_int(v):
                try: return int(float(str(v).replace(",",".")))
                except: return 0
            stj[ano] = {
                "cn": to_int(row.get("cn","")),
                "cp": to_int(row.get("cp","")),
                "tbaix": to_int(row.get("tbaix","")),
                "jg": to_int(row.get("jg",""))
            }

# Datajud API real (track_total_hits)
datajud = {
    2003:62, 2004:63, 2005:107, 2006:261, 2007:632, 2008:527, 2009:664, 2010:855,
    2011:3296, 2012:8073, 2013:14484, 2014:26112, 2015:34913, 2016:50270,
    2017:151977, 2018:294948, 2019:292525, 2020:310364, 2021:367585, 2022:367098,
    2023:381391, 2024:331998, 2025:37776
}

# CSVs locais
temos = {
    2005:107, 2006:261, 2007:632, 2008:527, 2009:664, 2010:855, 2011:3296,
    2012:8073, 2013:14484, 2014:26112, 2015:34913, 2016:50270,
    2017:148827, 2018:278702, 2019:275213, 2020:291850, 2021:333262,
    2022:335067, 2023:362335, 2024:299998, 2025:36955, 2026:144216
}

wb = Workbook()
hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(bold=True, color="FFFFFF", size=10)
alt = PatternFill("solid", fgColor="D6E4F0")
red = PatternFill("solid", fgColor="FFC7CE")
yellow = PatternFill("solid", fgColor="FFEB9C")
green = PatternFill("solid", fgColor="C6EFCE")
thin = Side(style="thin")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

# === ABA 1: CONTRASTE ===
ws = wb.active
ws.title = "Contraste Cobertura"
hdrs = ["Ano", "Casos Novos (JN)", "Datajud API", "% Datajud/JN", "Temos CSV",
        "% CSV/JN", "Gap JN-CSV", "% Gap", "Gap API-CSV", "Obs"]
widths = [8, 18, 14, 14, 14, 12, 14, 10, 14, 40]
for i, h in enumerate(hdrs, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd; c.alignment = Alignment(horizontal="center")
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w

r = 2
for ano in range(2005, 2027):
    jn = stj.get(ano, {}).get("cn", 0)
    api = datajud.get(ano, 0)
    csv_n = temos.get(ano, 0)

    pct_api = round(api/jn*100, 1) if jn > 0 else 0
    pct_csv = round(csv_n/jn*100, 1) if jn > 0 else 0
    gap_jn = jn - csv_n if jn > 0 else 0
    pct_gap = round(gap_jn/jn*100, 1) if jn > 0 else 0
    gap_api = api - csv_n

    if ano < 2009: obs = "Sem dados JN (pre-2009)"
    elif jn == 0 and ano <= 2010: obs = "JN=0 para este ano"
    elif ano == 2025: obs = "Ano em curso"
    elif ano == 2026: obs = "CKAN 144K com teor decisao"
    elif pct_csv >= 85: obs = "Cobertura boa"
    elif pct_csv >= 50: obs = "Cobertura parcial"
    elif pct_csv >= 10: obs = "Cobertura baixa"
    elif pct_csv > 0: obs = "Amostra minima"
    else: obs = ""

    vals = [ano,
            jn if jn > 0 else "-",
            api if api > 0 else "-",
            pct_api if jn > 0 else "-",
            csv_n,
            pct_csv if jn > 0 else "-",
            gap_jn if jn > 0 else "-",
            pct_gap if jn > 0 else "-",
            gap_api,
            obs]

    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=j+1, value=v)
        c.border = brd
        if j in (1,2,4,6,8) and isinstance(v, (int, float)):
            c.number_format = "#,##0"
        if j == 3 and isinstance(v, (int,float)):
            c.number_format = "0.0"
        if j == 5 and isinstance(v, (int,float)):
            c.number_format = "0.0"
            if v >= 85: c.fill = green
            elif v >= 50: c.fill = yellow
            else: c.fill = red
        if j == 7 and isinstance(v, (int,float)):
            c.number_format = "0.0"
            if v > 50: c.fill = red; c.font = Font(bold=True, color="CC0000")
            elif v > 15: c.fill = yellow
    r += 1

# Total
jn_tot = sum(stj.get(a,{}).get("cn",0) for a in range(2011,2025))
csv_tot = sum(temos.get(a,0) for a in range(2005,2027))
api_tot = sum(datajud.get(a,0) for a in range(2003,2026))
tf = PatternFill("solid", fgColor="FFC000")
for j, v in enumerate(["TOTAL", jn_tot, api_tot, "", csv_tot, "", jn_tot-csv_tot, "", api_tot-csv_tot, ""]):
    c = ws.cell(row=r, column=j+1, value=v)
    c.font = Font(bold=True); c.fill = tf; c.border = brd
    if j in (1,2,4,6,8) and isinstance(v, int): c.number_format = "#,##0"

# === ABA 2: PLANO ===
ws2 = wb.create_sheet("Plano de Acao")
plan = [
    ["LACUNA", "DIMENSAO", "SOLUCAO POSSIVEL", "PRIORIDADE"],
    ["2017-2024: 173K faltando nos CSVs",
     "Script paginava por semana, perdeu batches",
     "Re-extrair com paginacao por DIA (max ~1.6K/dia, nunca estoura 10K)",
     "ALTA"],
    ["3.193 sem dataAjuizamento",
     "Nunca capturados pelo filtro de data",
     "Query Datajud sem filtro de data, salvar em CSV separado",
     "ALTA"],
    ["2005-2010: <1K no Datajud vs 220-290K reais",
     "Datajud tem 0.05-0.37% do volume real",
     "CKAN STJ historico? API jurisprudencia STJ? Dados abertos STJ?",
     "MEDIA"],
    ["2011-2016: 3K-50K vs 260-334K reais",
     "Datajud tem 1-15% do real",
     "Mesmas fontes alternativas acima",
     "MEDIA"],
    ["2025: 37K vs ~478K esperado",
     "Ano em curso, Datajud atrasa",
     "Re-extrair + CKAN complementar",
     "BAIXA"],
]
for i, row in enumerate(plan):
    for j, v in enumerate(row):
        c = ws2.cell(row=i+1, column=j+1, value=v)
        c.border = brd
        if i == 0: c.font = hfont; c.fill = hfill
        if j == 3 and v == "ALTA": c.font = Font(bold=True, color="CC0000")
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 42
ws2.column_dimensions["C"].width = 58
ws2.column_dimensions["D"].width = 12

# === ABA 3: JN STJ ===
ws3 = wb.create_sheet("Justica em Numeros STJ")
for i, h in enumerate(["Ano","Casos Novos","Pendentes","Baixados","Julgados"], 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd
r3 = 2
for ano in sorted(stj.keys()):
    s = stj[ano]
    for j, v in enumerate([ano, s["cn"], s["cp"], s["tbaix"], s["jg"]]):
        c = ws3.cell(row=r3, column=j+1, value=v)
        c.border = brd; c.number_format = "#,##0"
        if r3 % 2 == 0: c.fill = alt
    r3 += 1
for i in range(1,6): ws3.column_dimensions[get_column_letter(i)].width = 16

# === ABA 4: NOTAS ===
ws4 = wb.create_sheet("Notas")
notas = [
    ("CONTRASTE DE COBERTURA STJ", True, 14, None),
    ("29/03/2026 | JudX | Claude Code", False, 10, "666666"),
    ("", False, 10, None),
    ("FONTES:", True, 12, None),
    ("  CNJ Justica em Numeros: dados-jn-23-set-2025.zip (2009-2024)", False, 11, None),
    ("  Datajud API: track_total_hits=true (total real: 3.370.504)", False, 11, None),
    ("  CKAN STJ: dadosabertos.web.stj.jus.br (2026)", False, 11, None),
    ("  CSVs locais: Desktop/backup_judx/resultados/stj_datajud/", False, 11, None),
    ("", False, 10, None),
    ("ACHADOS:", True, 12, None),
    ("  1. Datajud tem 3.370.504 processos STJ no indice", False, 11, None),
    ("  2. Extraimos 2.502.404 em CSV (74% do Datajud)", False, 11, None),
    ("  3. 173.453 faltam de 2017-2024 (paginacao incompleta)", False, 11, "CC0000"),
    ("  4. 3.193 sem data de ajuizamento (nunca capturados)", False, 11, "CC0000"),
    ("  5. Anos 2005-2016: Datajud tem <15% do real (JN/CNJ)", False, 11, None),
    ("  6. Volume real STJ: 295K-478K casos/ano (2011-2024)", False, 11, None),
    ("  7. O Datajud NAO e fonte completa para anos pre-2017", False, 11, None),
    ("", False, 10, None),
    ("IMPLICACAO PARA O PAPER:", True, 12, None),
    ("  Para 2017-2024: podemos completar via Datajud (re-extrair)", False, 11, None),
    ("  Para 2005-2016: Datajud e amostra, nao universo", False, 11, None),
    ("  Qualquer analise pre-2017 deve explicitar que usa amostra Datajud", False, 11, None),
    ("  JN/CNJ e a fonte autoritativa para volume agregado", False, 11, None),
]
for i, (txt, bold, sz, color) in enumerate(notas):
    c = ws4.cell(row=i+1, column=1, value=txt)
    c.font = Font(bold=bold, size=sz, color=color or "000000")
ws4.column_dimensions["A"].width = 75

wb.save(out)
print(f"Salvo: {out}")
print("4 abas: Contraste, Plano, JN STJ, Notas")
