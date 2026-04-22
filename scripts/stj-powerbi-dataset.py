"""stj-powerbi-dataset.py

Consolida os CSVs do dashboard STJ num único arquivo Excel estruturado em
modelo star-schema, pronto para ser importado no Power BI Desktop.

Entrada: CSVs em Desktop\backup_judx\resultados\2026-04-19_stj_dash_*.csv
Saída:   Desktop\backup_judx\resultados\JUDX_STJ_POWERBI.xlsx
         + JUDX_STJ_README_PowerBI.md (passo-a-passo em português)
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("C:/Users/medin/Desktop/backup_judx/resultados")
DATA = "2026-04-19"

def read_csv(path):
    rows = []
    with open(path, 'r', encoding='utf-8') as f:
        r = csv.reader(f)
        header = next(r)
        for row in r:
            rows.append(row)
    return header, rows

# === Estilos ===
font_header = Font(bold=True, color="FFFFFF", size=11)
fill_header = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
fill_kpi = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right")
thin = Side(border_style="thin", color="D0D7DE")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all

def autosize(ws):
    from openpyxl.cell.cell import MergedCell
    maxlens = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            letter = cell.column_letter
            val = str(cell.value) if cell.value is not None else ''
            if len(val) > maxlens.get(letter, 0):
                maxlens[letter] = len(val)
    for letter, w in maxlens.items():
        ws.column_dimensions[letter].width = min(w + 3, 60)

def aba_from_csv(wb, sheet_name, csv_name, title=None, notas=None):
    """Cria aba do arquivo Excel a partir de um CSV."""
    header, rows = read_csv(OUT / f"{DATA}_{csv_name}.csv")
    ws = wb.create_sheet(sheet_name)
    if title:
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color="1F6FEB")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        row_offset = 2
        if notas:
            ws.cell(row=2, column=1, value=notas).font = Font(italic=True, size=10, color="57606A")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(header))
            row_offset = 3
    else:
        row_offset = 1
    # cabeçalho
    for i, h in enumerate(header, 1):
        ws.cell(row=row_offset, column=i, value=h)
    style_header(ws, row=row_offset)
    # dados
    for r_idx, row in enumerate(rows, row_offset + 1):
        for c_idx, v in enumerate(row, 1):
            try: v_cast = int(v)
            except:
                try: v_cast = float(v)
                except: v_cast = v
            ws.cell(row=r_idx, column=c_idx, value=v_cast)
    autosize(ws)
    ws.freeze_panes = ws.cell(row=row_offset + 1, column=1)
    return ws

# ==========================================================
# Montagem do Excel
# ==========================================================
wb = Workbook()
# primeira aba = Resumo
ws0 = wb.active
ws0.title = "00_Resumo"

# KPIs principais (lê os CSVs pra computar)
_, classes = read_csv(OUT / f"{DATA}_stj_dash_classe.csv")
_, origens = read_csv(OUT / f"{DATA}_stj_dash_origem.csv")
_, segmentos = read_csv(OUT / f"{DATA}_stj_dash_segmento.csv")
_, resultados = read_csv(OUT / f"{DATA}_stj_dash_resultado.csv")
_, trilha = read_csv(OUT / f"{DATA}_stj_dash_posicao_trilha.csv")
_, interp = read_csv(OUT / f"{DATA}_stj_dash_interpenetracoes.csv")
_, relatores = read_csv(OUT / f"{DATA}_stj_dash_relator.csv")
_, assuntos = read_csv(OUT / f"{DATA}_stj_dash_assuntos.csv")

total_processos = sum(int(r[1]) for r in classes)
total_interpen = sum(int(r[2]) for r in interp)
n_classes = len(classes)
n_tribunais = len(origens)
n_relatores = len([r for r in relatores if not r[0].startswith('(')])

ws0['A1'] = "JudX · STJ — Universo Datajud"
ws0['A1'].font = Font(bold=True, size=18, color="1F6FEB")
ws0.merge_cells("A1:D1")

ws0['A2'] = "Pesquisa empírica Damares Medina · 19/abr/2026"
ws0['A2'].font = Font(italic=True, size=11, color="57606A")
ws0.merge_cells("A2:D2")

kpis = [
    ("Processos analisados",    total_processos,   "strings únicas (numeroProcesso)"),
    ("Classes distintas",        n_classes,          "tipos de recurso/ação (AREsp, HC, REsp...)"),
    ("Tribunais de origem",      n_tribunais,        "TJs, TRFs, STJ próprio, militares"),
    ("Ministros identificados",  n_relatores,        "gabinetes com relator extraível"),
    ("Interpenetrações",         total_interpen,     "processos cujo tipo atravessa ramos"),
]
ws0['A4'] = "Indicador";   ws0['B4'] = "Valor";   ws0['C4'] = "Descrição"
style_header(ws0, row=4)
for i, (lbl, val, desc) in enumerate(kpis, 5):
    ws0.cell(row=i, column=1, value=lbl).font = Font(bold=True)
    ws0.cell(row=i, column=2, value=val).alignment = align_right
    ws0.cell(row=i, column=2).fill = fill_kpi
    ws0.cell(row=i, column=2).number_format = '#,##0'
    ws0.cell(row=i, column=3, value=desc)

ws0['A12'] = "Abas do arquivo"
ws0['A12'].font = Font(bold=True, size=13, color="1F6FEB")
abas_info = [
    ("01_dim_classe",          "Tipos de ação/recurso no STJ"),
    ("02_dim_origem",          "Tribunais que alimentam o STJ (80 buckets)"),
    ("03_dim_segmento",        "Ramos de origem (Estadual, Federal, Superior, Militar)"),
    ("04_dim_resultado",       "Resultado final priorizado de cada processo"),
    ("05_dim_relator",         "Ministros (extraídos do nome do gabinete)"),
    ("06_dim_trilha",          "Posição da string na teoria dos Objetos Ancorados"),
    ("07_dim_assunto",         "Top 100 assuntos TPU-CNJ"),
    ("08_interpenetracoes",    "Classes que atravessam ramos (CC, Rcl, Pet, SLS)"),
    ("09_classe_x_origem_x_resultado", "Cruzamento triplo — o coração do dashboard"),
    ("10_relator_x_resultado", "Perfil decisório por ministro"),
]
ws0.cell(row=13, column=1, value="Aba"); ws0.cell(row=13, column=2, value="Conteúdo")
style_header(ws0, row=13)
for i, (aba, descr) in enumerate(abas_info, 14):
    ws0.cell(row=i, column=1, value=aba).font = Font(bold=True)
    ws0.cell(row=i, column=2, value=descr)

ws0['A26'] = "Como usar no Power BI Desktop"
ws0['A26'].font = Font(bold=True, size=13, color="1F6FEB")
passos = [
    "1. Abra o Power BI Desktop.",
    "2. Tela inicial → Obter Dados → Excel → escolha este arquivo.",
    "3. No Navegador, marque as abas 01 a 10 e clique em Carregar.",
    "4. No painel Modelo, o Power BI cria automaticamente as relações entre as dimensões e os cruzamentos.",
    "5. Arraste os campos para Visualizações (barras, treemap, matriz) conforme o recorte que quiser estudar.",
    "",
    "Atalho: dimensões (prefixo dim_) são filtros; cruzamentos (prefixo 09_ e 10_) são os fatos.",
    "",
    "Observação: a aba 05_dim_relator ainda usa o nome do gabinete como proxy do relator.",
    "A estrutura correta (ministro → Turma/Seção) será refeita espelhando o stf_master.",
]
for i, p in enumerate(passos, 27):
    ws0.cell(row=i, column=1, value=p)
    ws0.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
autosize(ws0)

# Dimensões e cruzamentos a partir dos CSVs
aba_from_csv(wb, "01_dim_classe",          "stj_dash_classe",
    "Tipos de ação no STJ", "Uma linha por tipo (AREsp, HC, REsp...). Total 59 distintos.")
aba_from_csv(wb, "02_dim_origem",          "stj_dash_origem",
    "Tribunais de origem", "Baseado nos dígitos 14-16 do CNJ (J.TR). 80 buckets.")
aba_from_csv(wb, "03_dim_segmento",        "stj_dash_segmento",
    "Ramos de origem", "Agregado por competência constitucional.")
aba_from_csv(wb, "04_dim_resultado",       "stj_dash_resultado",
    "Resultado final priorizado", "Classificação dos movimentos: mérito > não-conhecimento > prejudicado > desistência.")
aba_from_csv(wb, "05_dim_relator",         "stj_dash_relator",
    "Ministros (gabinetes)", "ATENÇÃO: gabinete ≠ órgão julgador. Campo a ser refeito espelhando stf_master.")
aba_from_csv(wb, "06_dim_trilha",          "stj_dash_posicao_trilha",
    "Posição na trilha", "Teoria dos Objetos Ancorados: brota_inadmissão, filha_direta, nó_interno, etc.")
aba_from_csv(wb, "07_dim_assunto",         "stj_dash_assuntos",
    "Assuntos TPU-CNJ (top 100)", "Primeiro assunto de cada processo, código do dicionário oficial CNJ.")
aba_from_csv(wb, "08_interpenetracoes",    "stj_dash_interpenetracoes",
    "Interpenetrações de ramo", "Classes que atravessam ramos: CC, Rcl, Pet, SLS, PU, SS.")
aba_from_csv(wb, "09_classe_x_origem_x_resultado", "stj_dash_crossab_classe_origem_resultado",
    "Cruzamento classe × origem × resultado", "O coração do dashboard: matriz tridimensional pronta para pivot.")
aba_from_csv(wb, "10_relator_x_resultado", "stj_dash_relator_x_resultado",
    "Perfil decisório por ministro", "Cruzamento ministro × resultado (mérito provido, não conhecido, etc).")

# Salva
excel_path = OUT / "JUDX_STJ_POWERBI.xlsx"
wb.save(excel_path)
print(f"[excel] {excel_path}")
print(f"  tamanho: {excel_path.stat().st_size/1024:.1f} KB")
print(f"  abas: {len(wb.sheetnames)}")

# ==========================================================
# README em português
# ==========================================================
readme = f"""# JUDX STJ — Dataset Power BI

Arquivo: `JUDX_STJ_POWERBI.xlsx`
Gerado em: 19/abr/2026
Universo: {total_processos:,} processos STJ (Datajud completo)

---

## Como abrir no Power BI Desktop

1. **Abra o Power BI Desktop** (se ainda não tiver, instale gratuito em powerbi.microsoft.com/desktop).
2. **Tela inicial** → botão **Obter Dados** → **Excel** → escolha `JUDX_STJ_POWERBI.xlsx`.
3. No **Navegador**, marque as dez abas (de `01_dim_classe` a `10_relator_x_resultado`) + a aba `00_Resumo` se quiser o overview.
4. Clique em **Carregar** (não precisa transformar — o formato já está limpo).
5. O Power BI cria automaticamente o modelo tabular. Se quiser, use **Modelo** (painel à esquerda) para ajustar relações entre as tabelas.

## Estrutura do arquivo (modelo star schema)

- **Aba `00_Resumo`** — visão geral do universo.
- **Abas `01_` a `08_`** (dimensões) — uma característica por processo: classe, origem, resultado, relator, assunto, etc.
- **Abas `09_` e `10_`** (fatos cruzados) — as tabelas que o Power BI usa como base dos gráficos.

## Visuais sugeridos (para começar)

- **Treemap**: aba `01_dim_classe` → campo `processo_curto` em Categoria + `ocorrencias` em Valores
- **Matriz**: aba `09_classe_x_origem_x_resultado` → Linhas = processo_curto, Colunas = resultado, Valores = ocorrencias
- **Mapa de árvore**: aba `10_relator_x_resultado` → Categoria = relator, Sub = resultado, Valor = ocorrencias
- **Gráfico de rosca**: aba `04_dim_resultado` → visualiza a proporção dos 9 resultados possíveis

## Limitação conhecida

A aba `05_dim_relator` ainda usa o nome do **gabinete** como proxy. No STJ (como no STF), **gabinete não é órgão julgador**: o relator é o ministro dono do gabinete, e o órgão julgador é a Turma/Seção/Corte Especial à qual ele pertence. Essa aba será refeita espelhando o padrão do `stf_master`.

## Arquivo fonte (caso queira importar direto)

Os 10 CSVs originais estão em `Desktop\\backup_judx\\resultados\\` com prefixo `2026-04-19_stj_dash_*.csv`. O Power BI também aceita CSVs diretamente (Obter Dados → Texto/CSV).
"""
readme_path = OUT / "JUDX_STJ_README_PowerBI.md"
readme_path.write_text(readme, encoding='utf-8')
print(f"[readme] {readme_path}")
print(f"  tamanho: {readme_path.stat().st_size/1024:.1f} KB")

print("\n[fim] Pacote Power BI pronto. Abra o arquivo .xlsx direto no Power BI Desktop.")
