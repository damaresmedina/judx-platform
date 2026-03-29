"""
Auditoria STJ final — 1 aba por ano com amostra + cobertura de campos
+ aba resumo geral + aba cobertura comparativa
"""
import csv, os, glob, re, random
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

dir_path = r"C:\Users\medin\Desktop\backup_judx\resultados\stj_datajud"
out_path = r"C:\Users\medin\Desktop\backup_judx\auditoria_stj_final_2026-03-29.xlsx"
files = sorted(glob.glob(os.path.join(dir_path, "*.csv")))

wb = Workbook()
hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(bold=True, color="FFFFFF", size=10)
alt_fill = PatternFill("solid", fgColor="D6E4F0")
red_font = Font(color="CC0000", bold=True)
orange_font = Font(color="FF8800")
green_font = Font(color="006600")
thin = Side(style="thin")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

def pct_font(v):
    if v < 50: return red_font
    if v < 80: return orange_font
    return Font(size=10)

def count_lines(fp):
    n = 0
    with open(fp, "rb") as f:
        for _ in f:
            n += 1
    return n - 1

# ========== Coleta por arquivo ==========
print("Processando...")
all_stats = []

for fp in files:
    fname = os.path.basename(fp)
    m = re.search(r"(\d{4})", fname)
    ano = int(m.group(1)) if m else 0
    is_ckan = "ckan" in fname
    fonte = "CKAN" if is_ckan else "Datajud"
    size_mb = round(os.path.getsize(fp) / (1024*1024), 1)
    total = count_lines(fp)

    # Campos a auditar
    if is_ckan:
        fields = ["numero_processo", "classe", "data_publicacao", "data_recebimento",
                  "data_distribuicao", "relator", "tipo_documento", "teor", "assuntos"]
    else:
        fields = ["numero_processo", "classe_nome", "data_ajuizamento", "relator",
                  "gabinete", "assuntos", "ultima_fase", "total_movimentos", "formato"]

    field_counts = {f: 0 for f in fields}
    classes = Counter()
    relatores = Counter()
    sample_rows = []
    all_buf = []

    with open(fp, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        real_fields = reader.fieldnames or []
        for idx, row in enumerate(reader):
            if idx < 3000:
                all_buf.append(row)
            if idx < 1000:
                for fld in fields:
                    val = (row.get(fld, "") or "").strip()
                    if val and val != "0":
                        field_counts[fld] += 1
                cl = (row.get("classe_nome", "") or row.get("classe", "") or "").strip()
                if cl: classes[cl] += 1
                rel = (row.get("relator", "") or "").strip()
                if rel: relatores[rel] += 1

    # Amostra: 30 aleatórios
    n_sample = min(30, len(all_buf))
    sample_rows = random.sample(all_buf, n_sample) if all_buf else []
    sampled = min(1000, len(all_buf))

    # Percentuais
    field_pcts = {}
    for fld in fields:
        field_pcts[fld] = round(field_counts[fld] / sampled * 100, 1) if sampled > 0 else 0

    all_stats.append({
        "ano": ano, "fonte": fonte, "fname": fname, "size_mb": size_mb,
        "total": total, "fields": fields, "field_pcts": field_pcts,
        "classes": classes, "relatores": relatores,
        "sample": sample_rows, "sampled": sampled,
        "real_fields": real_fields, "is_ckan": is_ckan
    })
    print(f"  {fname}: {total:>9,} processos | {size_mb} MB")

# ========== ABA POR ANO ==========
first = True
for st in all_stats:
    ano = st["ano"]
    fonte = st["fonte"]
    sheet_name = f"{ano} {fonte}"
    if first:
        ws = wb.active
        ws.title = sheet_name
        first = False
    else:
        ws = wb.create_sheet(sheet_name)

    # Bloco 1: Resumo do ano
    ws.cell(row=1, column=1, value=f"STJ {ano} ({fonte})").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Arquivo: {st['fname']}").font = Font(size=10, color="666666")
    ws.cell(row=2, column=4, value=f"{st['total']:,} processos").font = Font(bold=True, size=11)
    ws.cell(row=2, column=6, value=f"{st['size_mb']} MB").font = Font(size=10)

    # Bloco 2: Cobertura de campos
    ws.cell(row=4, column=1, value="COBERTURA DE CAMPOS").font = Font(bold=True, size=11)
    ws.cell(row=4, column=4, value=f"(amostra: {st['sampled']} linhas)").font = Font(italic=True, size=9, color="666666")

    r = 5
    for i, (fld, pct) in enumerate(st["field_pcts"].items()):
        ws.cell(row=r, column=1, value=fld).font = Font(bold=True, size=10)
        ws.cell(row=r, column=1).border = brd
        c = ws.cell(row=r, column=2, value=f"{pct}%")
        c.border = brd
        c.font = pct_font(pct)
        # Barra visual
        bar = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
        ws.cell(row=r, column=3, value=bar).font = Font(size=8, color="1F4E79")
        r += 1

    # Bloco 3: Top 10 classes
    r += 1
    ws.cell(row=r, column=1, value="TOP 10 CLASSES").font = Font(bold=True, size=11)
    r += 1
    for cl, cnt in st["classes"].most_common(10):
        ws.cell(row=r, column=1, value=cl).border = brd
        c = ws.cell(row=r, column=2, value=cnt)
        c.border = brd; c.number_format = "#,##0"
        pct = round(cnt / st["sampled"] * 100, 1) if st["sampled"] > 0 else 0
        ws.cell(row=r, column=3, value=f"{pct}%").border = brd
        r += 1

    # Bloco 4: Top 10 relatores
    r += 1
    ws.cell(row=r, column=1, value="TOP 10 RELATORES").font = Font(bold=True, size=11)
    r += 1
    for rel, cnt in st["relatores"].most_common(10):
        ws.cell(row=r, column=1, value=rel).border = brd
        c = ws.cell(row=r, column=2, value=cnt)
        c.border = brd; c.number_format = "#,##0"
        r += 1

    # Bloco 5: Amostra de registros
    r += 1
    ws.cell(row=r, column=1, value=f"AMOSTRA ({len(st['sample'])} registros aleatorios)").font = Font(bold=True, size=11)
    r += 1

    # Headers da amostra
    display_fields = st["fields"]
    for j, fld in enumerate(display_fields):
        c = ws.cell(row=r, column=j+1, value=fld)
        c.font = hfont; c.fill = hfill; c.border = brd
    r += 1

    for i, row in enumerate(st["sample"]):
        for j, fld in enumerate(display_fields):
            val = (row.get(fld, "") or "")
            # Truncar campos longos
            if len(val) > 120:
                val = val[:120] + "..."
            c = ws.cell(row=r, column=j+1, value=val)
            c.border = brd
            c.font = Font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if i % 2 == 1:
                c.fill = alt_fill
        r += 1

    # Ajustar larguras
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    for j in range(4, len(display_fields) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 25

# ========== ABA RESUMO GERAL ==========
ws_res = wb.create_sheet("Resumo Geral")

hdrs = ["Ano", "Fonte", "Processos", "MB", "% processo", "% classe", "% data",
        "% relator", "% assuntos", "% fase/teor", "Obs"]
widths_r = [8, 8, 14, 8, 10, 10, 10, 10, 12, 12, 30]
for i, h in enumerate(hdrs, 1):
    c = ws_res.cell(row=1, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd
for i, w in enumerate(widths_r):
    ws_res.column_dimensions[get_column_letter(i+1)].width = w

grand_total = 0
grand_mb = 0
r = 2
for st in all_stats:
    grand_total += st["total"]
    grand_mb += st["size_mb"]
    fp = st["field_pcts"]

    if st["is_ckan"]:
        pcts = [fp.get("numero_processo",0), fp.get("classe",0), fp.get("data_publicacao",0),
                fp.get("relator",0), fp.get("assuntos",0), fp.get("teor",0)]
    else:
        pcts = [fp.get("numero_processo",0), fp.get("classe_nome",0), fp.get("data_ajuizamento",0),
                fp.get("relator",0), fp.get("assuntos",0), fp.get("ultima_fase",0)]

    obs = ""
    if st["total"] < 500: obs = "Cobertura muito parcial"
    elif st["total"] < 5000: obs = "Cobertura parcial"

    vals = [st["ano"], st["fonte"], st["total"], st["size_mb"]] + pcts + [obs]
    for j, v in enumerate(vals):
        c = ws_res.cell(row=r, column=j+1, value=v)
        c.border = brd
        if j == 2: c.number_format = "#,##0"
        if j >= 4 and j <= 9 and isinstance(v, (int, float)):
            c.font = pct_font(v)
        if r % 2 == 0: c.fill = alt_fill
        if j == 10 and v: c.font = Font(color="CC0000", size=10)
    r += 1

# Total
tot_fill = PatternFill("solid", fgColor="FFC000")
for j, v in enumerate(["TOTAL", "", grand_total, round(grand_mb, 1)] + [""] * 7):
    c = ws_res.cell(row=r, column=j+1, value=v)
    c.font = Font(bold=True, size=11); c.fill = tot_fill; c.border = brd
    if j == 2: c.number_format = "#,##0"

# ========== ABA NOTAS ==========
ws_n = wb.create_sheet("Notas")
notas = [
    ("AUDITORIA FINAL — CORPUS STJ COMPLETO", True, 14, None),
    ("29/03/2026 | Claude Code (JudX)", False, 10, "666666"),
    ("", False, 10, None),
    (f"CORPUS: {grand_total:,} processos | {round(grand_mb,1)} MB | 22 anos (2005-2026)", True, 12, "1F4E79"),
    ("", False, 10, None),
    ("FONTES:", True, 12, None),
    ("  Datajud CNJ: 2005-2025 (api-publica.datajud.cnj.jus.br)", False, 11, None),
    ("  CKAN STJ: 2026 (dadosabertos.web.stj.jus.br)", False, 11, None),
    ("", False, 10, None),
    ("PARSER RELATOR (atualizado 29/mar):", True, 12, None),
    ("  Ministro/Ministra: extraido do campo gabinete", False, 11, None),
    ("  Desembargador convocado: marcado como DESEMB. CONV. NOME", False, 11, None),
    ("  Presidencia: marcado como PRESIDENTE STJ", False, 11, None),
    ("  Vice-Presidencia: marcado como VICE-PRESIDENTE STJ", False, 11, None),
    ("  Cobertura relator: 69.7% -> 99.8% apos reparse", False, 11, "006600"),
    ("", False, 10, None),
    ("COBERTURA POR PERIODO:", True, 12, None),
    ("  2005-2010: <1.000/ano — Datajud retroativo parcial", False, 11, "CC0000"),
    ("  2011-2013: 3K-14K/ano — crescente", False, 11, "FF8800"),
    ("  2014-2016: 26K-50K/ano — robusta", False, 11, None),
    ("  2017-2024: 148K-362K/ano — massiva", False, 11, "006600"),
    ("  2025: 37K (pode estar incompleto)", False, 11, "FF8800"),
    ("  2026: 144K via CKAN (unico com teor da decisao)", False, 11, None),
    ("", False, 10, None),
    ("CADA ABA:", True, 12, None),
    ("  1. Cobertura de campos com barra visual", False, 11, None),
    ("  2. Top 10 classes processuais", False, 11, None),
    ("  3. Top 10 relatores/ministros", False, 11, None),
    ("  4. 30 registros aleatorios com todos os campos", False, 11, None),
    ("", False, 10, None),
    ("LOCAL DOS CSVs: Desktop/backup_judx/resultados/stj_datajud/", False, 11, None),
]
for i, (txt, bold, sz, color) in enumerate(notas):
    c = ws_n.cell(row=i+1, column=1, value=txt)
    c.font = Font(bold=bold, size=sz, color=color or "000000")
ws_n.column_dimensions["A"].width = 75

# Mover Resumo e Notas para o início
wb.move_sheet("Resumo Geral", offset=-len(all_stats))
wb.move_sheet("Notas", offset=-len(all_stats))

wb.save(out_path)
print(f"\nSalvo: {out_path}")
print(f"{len(all_stats)+2} abas: Notas + Resumo + 1 por ano/fonte")
print(f"Total: {grand_total:,} processos | {round(grand_mb,1)} MB")
