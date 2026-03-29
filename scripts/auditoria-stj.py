"""
Auditoria STJ — amostragem rápida (1000 linhas por arquivo) + contagem total
"""
import csv, os, glob, re, subprocess
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

dir_path = r'C:\Users\medin\Desktop\backup_judx\resultados\stj_datajud'
out_path = r'C:\Users\medin\Desktop\backup_judx\auditoria_stj_datajud_2026-03-29.xlsx'
files = sorted(glob.glob(os.path.join(dir_path, '*.csv')))

SAMPLE = 1000  # linhas amostradas por arquivo

wb = Workbook()
hfill = PatternFill('solid', fgColor='1F4E79')
hfont = Font(bold=True, color='FFFFFF', size=11)
alt_fill = PatternFill('solid', fgColor='D6E4F0')
thin = Side(style='thin')
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_headers(ws, headers, widths):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = Alignment(horizontal='center'); c.border = brd
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def count_lines(fp):
    """Conta linhas rápido sem ler conteúdo"""
    count = 0
    with open(fp, 'rb') as f:
        for _ in f:
            count += 1
    return count - 1  # menos header

def sample_file(fp, n=SAMPLE):
    """Lê header + primeiras n linhas"""
    classes = Counter()
    relatores = set()
    com_rel = 0; sem_rel = 0
    datas = []
    read = 0
    with open(fp, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if read >= n: break
            cl = row.get('classe_nome', '').strip()
            if cl: classes[cl] += 1
            rel = row.get('relator', '').strip()
            if rel: relatores.add(rel); com_rel += 1
            else: sem_rel += 1
            d = row.get('data_ajuizamento', '') or row.get('data_publicacao', '')
            if d: datas.append(d)
            read += 1
    datas.sort()
    pct = round(com_rel/(com_rel+sem_rel)*100, 1) if (com_rel+sem_rel) > 0 else 0
    return {
        'classes_top': classes.most_common(5),
        'n_classes': len(classes),
        'n_relatores': len(relatores),
        'com_rel': com_rel, 'sem_rel': sem_rel, 'pct': pct,
        'primeiro': datas[0] if datas else '-',
        'ultimo': datas[-1] if datas else '-',
        'amostrado': read
    }

# === COLETA ===
print("Contando linhas e amostrando...")
file_data = []
for fp in files:
    fname = os.path.basename(fp)
    size_mb = round(os.path.getsize(fp) / (1024*1024), 1)
    fonte = 'CKAN STJ' if 'ckan' in fname else 'Datajud CNJ'
    m = re.search(r'(\d{4})', fname)
    ano = int(m.group(1)) if m else 0

    total = count_lines(fp)
    s = sample_file(fp)
    print(f"  {fname}: {total:,} proc, {s['n_classes']} classes, {s['pct']}% relator")

    file_data.append({
        'fname': fname, 'size_mb': size_mb, 'ano': ano, 'fonte': fonte,
        'processos': total, **s
    })

# === ABA 1: INVENTÁRIO ===
ws = wb.active; ws.title = 'Inventario'
hdrs = ['Arquivo','MB','Processos','Primeiro*','Ultimo*','Classes*','Relatores*','% Relator*','Fonte','Amostra']
set_headers(ws, hdrs, [30, 8, 14, 16, 16, 10, 12, 12, 12, 10])

grand = sum(d['processos'] for d in file_data)
grand_mb = sum(d['size_mb'] for d in file_data)

for i, d in enumerate(file_data):
    r = i + 2
    vals = [d['fname'], d['size_mb'], d['processos'], d['primeiro'], d['ultimo'],
            d['n_classes'], d['n_relatores'], d['pct'], d['fonte'], d['amostrado']]
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=j+1, value=v)
        c.border = brd
        if i % 2 == 1: c.fill = alt_fill
        if j == 2: c.number_format = '#,##0'

# Total
r = len(file_data) + 2
tot_fill = PatternFill('solid', fgColor='FFC000')
for j, v in enumerate(['TOTAL', round(grand_mb,1), grand, '', '', '', '', '', '', '']):
    c = ws.cell(row=r, column=j+1, value=v)
    c.font = Font(bold=True); c.fill = tot_fill; c.border = brd
    if j == 2: c.number_format = '#,##0'

# Nota
r2 = len(file_data) + 4
ws.cell(row=r2, column=1, value='* Campos marcados com asterisco baseados em amostra de 1.000 linhas por arquivo').font = Font(italic=True, color='666666')

# === ABA 2: SÉRIE TEMPORAL ===
ws2 = wb.create_sheet('Serie Temporal')
set_headers(ws2, ['Ano','Fonte','Processos','Classes*','Relatores*','% Relator*','Obs'], [8,12,14,10,12,12,30])

for i, d in enumerate(file_data):
    r = i + 2
    obs = ''
    if d['processos'] < 500: obs = 'Cobertura muito parcial'
    elif d['processos'] < 5000: obs = 'Cobertura parcial'
    elif d['pct'] < 60: obs = 'Muitos sem relator'
    for j, v in enumerate([d['ano'], d['fonte'], d['processos'], d['n_classes'], d['n_relatores'], d['pct'], obs]):
        c = ws2.cell(row=r, column=j+1, value=v)
        c.border = brd
        if i % 2 == 1: c.fill = alt_fill
        if j == 2: c.number_format = '#,##0'
        if j == 6 and v: c.font = Font(color='CC0000')

# === ABA 3: CLASSES TOP POR ANO ===
ws3 = wb.create_sheet('Classes Top por Ano')
set_headers(ws3, ['Ano','Classe','Qtd na Amostra','Rank'], [8, 45, 16, 8])
r3 = 2
for d in file_data:
    for rank, (cl, cnt) in enumerate(d['classes_top'], 1):
        for j, v in enumerate([d['ano'], cl, cnt, rank]):
            c = ws3.cell(row=r3, column=j+1, value=v)
            c.border = brd
            if r3 % 2 == 0: c.fill = alt_fill
            if j == 2: c.number_format = '#,##0'
        r3 += 1

# === ABA 4: NOTAS ===
ws4 = wb.create_sheet('Notas')
notas = [
    ('AUDITORIA STJ — CORPUS DATAJUD + CKAN', True, 14),
    ('Data: 29/03/2026 | Gerado por: Claude Code (JudX)', False, 11),
    ('', False, 11),
    ('FONTES:', True, 11),
    ('  Datajud CNJ (api-publica.datajud.cnj.jus.br): anos 2005-2025', False, 11),
    ('  CKAN STJ (dadosabertos.web.stj.jus.br): ano 2026', False, 11),
    ('', False, 11),
    ('METODOLOGIA:', True, 11),
    ('  Contagem total: todas as linhas de cada CSV (exata)', False, 11),
    ('  Classes, relatores, datas: amostra das primeiras 1.000 linhas', False, 11),
    ('  Campos com * na planilha sao baseados em amostra', False, 11),
    ('', False, 11),
    ('LIMITACOES:', True, 11),
    ('  Datajud tem cobertura retroativa parcial antes de 2014', False, 11),
    ('  Anos 2005-2010: <1.000 processos cada — amostra, nao universo', False, 11),
    ('  2025 pode estar incompleto (37K vs ~300K em anos anteriores)', False, 11),
    ('  2026 Datajud vazio; dados 2026 vem do CKAN STJ', False, 11),
    ('', False, 11),
    (f'CORPUS TOTAL: {grand:,} processos', True, 13),
    ('PERIODO: 2005-2026 (22 anos)', True, 12),
    (f'TAMANHO TOTAL: {round(grand_mb,1)} MB em CSV', True, 12),
    ('LOCAL: Desktop/backup_judx/resultados/stj_datajud/', False, 11),
]
for i, (txt, bold, sz) in enumerate(notas):
    c = ws4.cell(row=i+1, column=1, value=txt)
    c.font = Font(bold=bold, size=sz)
ws4.column_dimensions['A'].width = 70

wb.save(out_path)
print(f"\nSalvo: {out_path}")
print(f"4 abas | {grand:,} processos | {round(grand_mb,1)} MB")
