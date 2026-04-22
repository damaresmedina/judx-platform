"""Gera tabela Excel normalizada dos 174 ministros do STF.

Fonte primária: C:/Users/medin/Downloads/STF ministros 2026.xlsx
Complementos (suprem lacunas): excel_parsed_174.json e stf_todos_ministros_consolidado.json

Ordem: mais antigo → mais novo (antiguidade crescente)
Datas normalizadas: 'Decreto de 31 de janeiro de 2024' → '2024-01-31'
Texto normalizado: sem acento, UPPERCASE nos nomes, sem HTML entities, sem espaços múltiplos.
"""
import re, json, unicodedata, html
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC_XLSX = "C:/Users/medin/Downloads/STF ministros 2026.xlsx"
JSON_174 = "C:/Users/medin/Desktop/stf no diva/dados_ministros/excel_parsed_174.json"
JSON_BIO = "C:/Users/medin/Desktop/backup_judx/resultados/stf_todos_ministros_consolidado.json"
OUT = Path("C:/Users/medin/Desktop/backup_judx/resultados/stf_172_ministros_FINAL.xlsx")

MESES = {'janeiro':1,'fevereiro':2,'março':3,'marco':3,'abril':4,'maio':5,'junho':6,
         'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12,
         'jan':1,'fev':2,'mar':3,'abr':4,'mai':5,'jun':6,'jul':7,'ago':8,'set':9,'out':10,'nov':11,'dez':12}

def normalizar_data(s):
    """Converte várias formas de data para YYYY-MM-DD."""
    if s is None or s == '': return ''
    s = str(s).strip()
    # já ISO
    m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
    if m: return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # DD/MM/YYYY
    m = re.match(r'^(\d{1,2})[/](\d{1,2})[/](\d{4})', s)
    if m: return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    # "Decreto de DD de MÊS de YYYY" ou "DD de MÊS de YYYY"
    m = re.search(r'(\d{1,2})\s+de\s+([a-zçA-Z]+)\s+de\s+(\d{4})', s)
    if m:
        d = int(m.group(1)); mes_s = m.group(2).lower(); y = int(m.group(3))
        mes_s_norm = ''.join(c for c in unicodedata.normalize('NFKD', mes_s) if not unicodedata.combining(c))
        mes = MESES.get(mes_s) or MESES.get(mes_s_norm)
        if mes: return f"{y:04d}-{mes:02d}-{d:02d}"
    # só "MÊS de YYYY"
    m = re.search(r'([a-zçA-Z]+)\s+de\s+(\d{4})', s)
    if m:
        mes_s_norm = ''.join(c for c in unicodedata.normalize('NFKD', m.group(1).lower()) if not unicodedata.combining(c))
        mes = MESES.get(mes_s_norm)
        y = int(m.group(2))
        if mes: return f"{y:04d}-{mes:02d}-01"
    # só ano "1890" ou "apenas ano"
    m = re.fullmatch(r'(\d{4})', s)
    if m: return f"{m.group(1)}-01-01"
    # se tem apenas dia + mês + ano colados sem padrão, devolve texto bruto
    return s

def limpar(s):
    """Remove HTML entities, aspas tipográficas, acentos, colapsa espaços."""
    if s is None: return ''
    s = str(s)
    s = html.unescape(s)
    s = s.replace("’","'").replace("‘","'").replace("“","\"").replace("”","\"")
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def limpar_upper(s):
    return limpar(s).upper()

def norm_str(s):
    if s is None: return ''
    return str(s).strip()

# ============================================================
# 1. Carrega Excel fonte
# ============================================================
wb = load_workbook(SRC_XLSX, data_only=True)
ws = wb['Planilha1']

# Mapeia cabeçalho
headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column+1) if ws.cell(row=1, column=c).value}
def col(name):
    # match por substring (case-insensitive) se não bater direto
    if name in headers: return headers[name]
    for h, c in headers.items():
        if h and name.lower() in str(h).lower(): return c
    return None

C_NOME     = col('Nome do Ministro')
C_IND_PRES = col('Indicação Presidencial')
C_LOCAL_IND= col('local of birth presidente')
C_NOMEACAO = col('Nomeação')
C_PAI      = col('Pai')
C_NASCIM   = col('DATE OF BIRTH')
C_LOCAL    = col('LOCAL OF BIRTH')
C_CARREIRA = col('CARRER')
C_FAC      = col('Faculdade')
C_POSSE    = col('Posse')
C_SAIDA_T  = col('Aposentadoria')
C_SAIDA_DT = col('Dt Saída')
C_ANTE     = col('Antecessor')
C_SUCE     = col('Sucessor')

ministros_xls = []
for r in range(2, ws.max_row+1):
    nome = norm_str(ws.cell(row=r, column=C_NOME).value)
    if not nome or len(nome) > 100: continue
    if nome.lower().startswith('a faculdade'): continue
    ministros_xls.append({
        'nome': nome,
        'indicacao_presidente': norm_str(ws.cell(row=r, column=C_IND_PRES).value) if C_IND_PRES else '',
        'local_presidente': norm_str(ws.cell(row=r, column=C_LOCAL_IND).value) if C_LOCAL_IND else '',
        'nomeacao_raw': norm_str(ws.cell(row=r, column=C_NOMEACAO).value) if C_NOMEACAO else '',
        'pai': norm_str(ws.cell(row=r, column=C_PAI).value) if C_PAI else '',
        'nascimento_raw': norm_str(ws.cell(row=r, column=C_NASCIM).value) if C_NASCIM else '',
        'local_nascimento': norm_str(ws.cell(row=r, column=C_LOCAL).value) if C_LOCAL else '',
        'carreira': norm_str(ws.cell(row=r, column=C_CARREIRA).value) if C_CARREIRA else '',
        'faculdade': norm_str(ws.cell(row=r, column=C_FAC).value) if C_FAC else '',
        'posse_raw': norm_str(ws.cell(row=r, column=C_POSSE).value) if C_POSSE else '',
        'saida_tipo': norm_str(ws.cell(row=r, column=C_SAIDA_T).value) if C_SAIDA_T else '',
        'saida_data_raw': norm_str(ws.cell(row=r, column=C_SAIDA_DT).value) if C_SAIDA_DT else '',
        'antecessor': norm_str(ws.cell(row=r, column=C_ANTE).value) if C_ANTE else '',
        'sucessor': norm_str(ws.cell(row=r, column=C_SUCE).value) if C_SUCE else '',
        '_linha_excel': r,
    })

print(f"ministros lidos do xlsx: {len(ministros_xls)}")

# ============================================================
# 2. Carrega JSONs de complemento
# ============================================================
with open(JSON_174, encoding='utf-8') as f:
    json174 = json.load(f)
with open(JSON_BIO, encoding='utf-8') as f:
    json_bio = json.load(f)

def sa(s):
    return ''.join(c for c in unicodedata.normalize('NFKD', str(s or '')) if not unicodedata.combining(c)).lower().strip()

json174_idx = {sa(m.get('nome','')): m for m in json174 if m.get('nome')}
bio_idx = {}
for m in json_bio:
    for k in (m.get('nome'), m.get('nome_completo')):
        if k: bio_idx[sa(k)] = m

def dmY(s):
    if not s: return ''
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', str(s).strip())
    return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}" if m else ''

# ============================================================
# 3. Normaliza + supre lacunas
# ============================================================
linhas = []
for m in ministros_xls:
    key = sa(m['nome'])
    j174 = json174_idx.get(key, {})
    bio  = bio_idx.get(key, {})

    # Nascimento
    nascimento = normalizar_data(m['nascimento_raw']) or dmY(bio.get('nascimento','').split(',')[0] if bio.get('nascimento') else '') or j174.get('nascimento_data','') or ''
    # Local nascimento
    local_nasc = m['local_nascimento'] or (bio.get('nascimento','').split(',',1)[1].strip() if ',' in (bio.get('nascimento','') or '') else '') or j174.get('nascimento_local','') or ''
    # Nomeação
    nomeacao = normalizar_data(m['nomeacao_raw']) or dmY(bio.get('nomeacao',''))
    # Posse
    posse = normalizar_data(m['posse_raw']) or dmY(bio.get('posse_stf','')) or j174.get('posse_stf_data','') or ''
    # Saída
    saida_tipo = m['saida_tipo'] or j174.get('saida_tipo','')
    saida_data = normalizar_data(m['saida_data_raw']) or dmY(bio.get('aposentadoria','')) or j174.get('saida_data','') or ''
    # Carreira
    carreira = m['carreira'] or j174.get('carreira','') or ''
    # Faculdade
    faculdade = m['faculdade'] or j174.get('faculdade','') or ''
    # Pai
    pai = m['pai'] or j174.get('pai','') or ''
    # Indicação
    ind_pres = m['indicacao_presidente'] or j174.get('presidente_indicou','') or ''
    local_pres = m['local_presidente']
    # Antecessor / Sucessor
    antecessor = m['antecessor'] or j174.get('antecessor','') or ''
    sucessor   = m['sucessor']   or j174.get('sucessor','')   or ''
    # Gênero, idade na posse
    genero = j174.get('genero','') or ''
    idade_posse = j174.get('idade_posse','') or ''

    linhas.append({
        'nome': limpar_upper(m['nome']),
        'nascimento': nascimento,
        'local_nascimento': limpar(local_nasc),
        'faculdade': limpar(faculdade),
        'carreira': limpar(carreira),
        'pai': limpar(pai),
        'indicacao_presidente': limpar(ind_pres),
        'local_presidente': limpar(local_pres),
        'nomeacao': nomeacao,
        'posse': posse,
        'idade_posse': limpar(idade_posse),
        'genero': limpar(genero).upper(),
        'saida_tipo': limpar(saida_tipo).upper(),
        'saida_data': saida_data,
        'antecessor': limpar_upper(antecessor),
        'sucessor': limpar_upper(sucessor),
    })

# Consolida Rezek (2 passagens) em 1 linha com coluna 'passagens'
from collections import defaultdict
grupos = defaultdict(list)
for r in linhas: grupos[r['nome']].append(r)
linhas_u = []
for nome, regs in grupos.items():
    if len(regs) == 1:
        regs[0]['passagens'] = ''
        linhas_u.append(regs[0])
    else:
        regs.sort(key=lambda x: x['posse'] or '9999')
        p = regs[0]
        p['passagens'] = ' | '.join(f"{r['posse']} -> {r['saida_data'] or 'atual'}" for r in regs)
        p['saida_data'] = regs[-1]['saida_data']
        p['saida_tipo'] = regs[-1]['saida_tipo']
        linhas_u.append(p)
linhas = linhas_u

# Ordem decrescente de antiguidade (convenção STF): #1 = mais recente, #172 = mais antigo
def sortkey(r):
    return (r['posse'] or '0000-00-00', r['nome'])
linhas.sort(key=sortkey, reverse=True)
for i, r in enumerate(linhas, 1):
    r['antiguidade'] = i

# ============================================================
# 4. Escreve Excel
# ============================================================
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "174 Ministros STF"

cols = ['antiguidade','nome','nascimento','local_nascimento','faculdade','carreira','pai',
        'indicacao_presidente','local_presidente','nomeacao','posse','idade_posse','genero',
        'saida_tipo','saida_data','passagens','antecessor','sucessor']

HEADERS = {
    'antiguidade':'#','nome':'NOME','nascimento':'NASCIMENTO','local_nascimento':'LOCAL NASCIMENTO',
    'faculdade':'FACULDADE','carreira':'CARREIRA','pai':'PAI',
    'indicacao_presidente':'INDICADO POR','local_presidente':'LOCAL DO PRESIDENTE',
    'nomeacao':'NOMEACAO','posse':'POSSE','idade_posse':'IDADE POSSE','genero':'GENERO',
    'saida_tipo':'SAIDA (TIPO)','saida_data':'SAIDA (DATA)','passagens':'PASSAGENS',
    'antecessor':'ANTECESSOR','sucessor':'SUCESSOR',
}
WIDTHS = {
    'antiguidade':5,'nome':45,'nascimento':12,'local_nascimento':22,'faculdade':28,'carreira':40,'pai':30,
    'indicacao_presidente':32,'local_presidente':18,'nomeacao':12,'posse':12,'idade_posse':7,'genero':7,
    'saida_tipo':20,'saida_data':12,'passagens':38,'antecessor':35,'sucessor':35,
}
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(bold=True, color='FFFFFF')
for idx, c in enumerate(cols, 1):
    cell = ws_out.cell(row=1, column=idx, value=HEADERS[c])
    cell.fill = header_fill; cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_out.column_dimensions[get_column_letter(idx)].width = WIDTHS[c]
ws_out.row_dimensions[1].height = 32

for i, r in enumerate(linhas, 2):
    for j, c in enumerate(cols, 1):
        cell = ws_out.cell(row=i, column=j, value=r.get(c,''))
        if c in ('antiguidade','nascimento','nomeacao','posse','saida_data','idade_posse','genero'):
            cell.alignment = Alignment(horizontal='center')

ws_out.freeze_panes = 'C2'
wb_out.save(OUT)

print(f"[ok] {OUT}")
print(f"linhas: {len(linhas)}")
# estatísticas de preenchimento
for c in ['nascimento','nomeacao','posse','faculdade','carreira','antecessor','sucessor']:
    n = sum(1 for r in linhas if r[c])
    print(f"  {c}: {n}/{len(linhas)} preenchidos")
