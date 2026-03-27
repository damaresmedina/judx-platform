"""
inventario-completo.py — Auditoria e inventário de tudo que existe
Gera Excel com múltiplas abas:
1. Inventário de arquivos locais
2. Estado do banco JudX (todas as tabelas com tamanhos)
3. Estado do banco ICONS
4. STF consolidado por processo (decisões + partes + ambiente)
5. STF por ambiente de julgamento
6. STJ consolidado por tema
7. STJ processos-semente com tribunal
8. Resumo executivo

Salva em Desktop/backup_judx/INVENTARIO_COMPLETO_YYYY-MM-DD.xlsx
"""
import sys, os, glob
sys.stdout.reconfigure(encoding='utf-8')

import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from datetime import datetime

JUDX = 'postgresql://postgres:Zb9cHoRww7WxgT0C@db.ejwyguskoiraredinqmb.supabase.co:5432/postgres'
ICONS = 'postgresql://postgres:RHuQvsf4shpsPRjP@db.hetuhkhhppxjliiaerlu.supabase.co:6543/postgres'
OUT_DIR = r'C:\Users\medin\Desktop\backup_judx'
os.makedirs(OUT_DIR, exist_ok=True)

NOW = datetime.now()
OUT = os.path.join(OUT_DIR, f'INVENTARIO_COMPLETO_{NOW.strftime("%Y-%m-%d")}.xlsx')

# Styles
HF = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
HB = PatternFill(start_color='1A2744', end_color='1A2744', fill_type='solid')
SF = Font(name='Calibri', bold=True, size=10, color='B8860B')  # gold section
SB = PatternFill(start_color='F5F0E8', end_color='F5F0E8', fill_type='solid')
DF = Font(name='Calibri', size=9)
NF = Font(name='Consolas', size=9)

def make_header(ws, headers, row=1):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HF; c.fill = HB; c.alignment = Alignment(horizontal='center')

def write_rows(ws, rows, start_row=2):
    for ri, row in enumerate(rows, start_row):
        for ci, val in enumerate(row, 1):
            if isinstance(val, (dict, list)):
                val = str(val)[:300]
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = DF
    return start_row + len(rows)

def auto_width(ws, headers):
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max(len(str(h))+4, 12), 55)
    ws.freeze_panes = 'A2'

def query(conn, sql):
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    cur.close()
    return rows

print(f'Gerando inventário completo...')
print(f'Output: {OUT}')

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════
# 1. ARQUIVOS LOCAIS
# ══════════════════════════════════════════════════════════
print('\n1. Mapeando arquivos locais...')
ws = wb.create_sheet('Arquivos Locais')
headers = ['Diretório', 'Arquivo', 'Tipo', 'Tamanho (KB)', 'Modificado']
make_header(ws, headers)

search_dirs = [
    r'C:\Users\medin\projetos\judx-platform',
    r'C:\Users\medin\projetos\judx-platform\scripts',
    r'C:\Users\medin\projetos\judx-platform\logs',
    r'C:\Users\medin\projetos\judx-platform\skills',
    r'C:\projetos\judx',
    r'C:\projetos\judx\STF',
    r'C:\projetos\icons',
    r'C:\projetos\icons-cartografia',
    r'C:\Users\medin\Desktop\backup_judx',
    r'C:\Users\medin\Desktop\bicocca milano',
    r'C:\Users\medin\Desktop\infoprodutos',
]

ri = 2
for d in search_dirs:
    if not os.path.exists(d):
        continue
    for f in sorted(os.listdir(d)):
        fp = os.path.join(d, f)
        if os.path.isfile(fp):
            ext = os.path.splitext(f)[1].lower()
            if ext in ['.mjs','.js','.py','.ts','.sql','.md','.html','.json','.csv','.xlsx','.xls','.docx','.pdf','.txt','.log']:
                sz = os.path.getsize(fp) / 1024
                mod = datetime.fromtimestamp(os.path.getmtime(fp)).strftime('%Y-%m-%d %H:%M')
                ws.cell(row=ri, column=1, value=d).font = DF
                ws.cell(row=ri, column=2, value=f).font = DF
                ws.cell(row=ri, column=3, value=ext).font = DF
                ws.cell(row=ri, column=4, value=round(sz, 1)).font = NF
                ws.cell(row=ri, column=5, value=mod).font = DF
                ri += 1

auto_width(ws, headers)
print(f'  {ri-2} arquivos mapeados')

# ══════════════════════════════════════════════════════════
# 2. BANCO JUDX — INVENTÁRIO DE TABELAS
# ══════════════════════════════════════════════════════════
print('\n2. Inventário banco JudX...')
ws = wb.create_sheet('Banco JudX')

conn = psycopg2.connect(JUDX, sslmode='require')

# All public tables with row counts
tables_sql = """
SELECT t.table_name,
  (SELECT COUNT(*) FROM information_schema.columns c WHERE c.table_name = t.table_name AND c.table_schema='public') as colunas
FROM information_schema.tables t
WHERE t.table_schema = 'public' AND t.table_type = 'BASE TABLE'
ORDER BY t.table_name
"""
tables = query(conn, tables_sql)

headers = ['Tabela', 'Colunas', 'Registros', 'Último registro']
make_header(ws, headers)

ri = 2
for table_name, ncols in tables:
    try:
        cnt = query(conn, f"SELECT COUNT(*) FROM {table_name}")[0][0]
    except:
        cnt = 'ERRO'
        conn.rollback()
    try:
        last = query(conn, f"SELECT MAX(created_at)::text FROM {table_name}")[0][0]
    except:
        last = None
        conn.rollback()

    ws.cell(row=ri, column=1, value=table_name).font = DF
    ws.cell(row=ri, column=2, value=ncols).font = NF
    ws.cell(row=ri, column=3, value=cnt).font = NF
    ws.cell(row=ri, column=4, value=str(last)[:19] if last else '').font = DF
    ri += 1

auto_width(ws, headers)
print(f'  {len(tables)} tabelas')

# ══════════════════════════════════════════════════════════
# 3. BANCO ICONS
# ══════════════════════════════════════════════════════════
print('\n3. Inventário banco ICONS...')
ws = wb.create_sheet('Banco ICONS')

conn_icons = psycopg2.connect(ICONS, sslmode='require')
headers = ['Tabela/Tipo', 'Registros']
make_header(ws, headers)

icons_queries = [
    ('objects (total)', "SELECT COUNT(*) FROM objects"),
    ('objects: registro_jurisprudencial', "SELECT COUNT(*) FROM objects WHERE type_slug='registro_jurisprudencial'"),
    ('objects: processo', "SELECT COUNT(*) FROM objects WHERE type_slug='processo'"),
    ('objects: artigo', "SELECT COUNT(*) FROM objects WHERE type_slug='artigo'"),
    ('objects: inciso', "SELECT COUNT(*) FROM objects WHERE type_slug='inciso'"),
    ('objects: paragrafo', "SELECT COUNT(*) FROM objects WHERE type_slug='paragrafo'"),
    ('objects: alinea', "SELECT COUNT(*) FROM objects WHERE type_slug='alinea'"),
    ('objects: tema_repetitivo_stj', "SELECT COUNT(*) FROM objects WHERE type_slug='tema_repetitivo_stj'"),
    ('edges (total)', "SELECT COUNT(*) FROM edges"),
    ('edges: ancora_normativa', "SELECT COUNT(*) FROM edges WHERE type_slug='ancora_normativa'"),
    ('edges: ancora_processual', "SELECT COUNT(*) FROM edges WHERE type_slug='ancora_processual'"),
    ('edges: relator_de', "SELECT COUNT(*) FROM edges WHERE type_slug='relator_de'"),
    ('edges: produzido_por', "SELECT COUNT(*) FROM edges WHERE type_slug='produzido_por'"),
]

ri = 2
for label, sql in icons_queries:
    try:
        val = query(conn_icons, sql)[0][0]
    except:
        val = 'ERRO'
        conn_icons.rollback()
    ws.cell(row=ri, column=1, value=label).font = DF
    ws.cell(row=ri, column=2, value=val).font = NF
    ri += 1

auto_width(ws, headers)
conn_icons.close()

# ══════════════════════════════════════════════════════════
# 4. STF CONSOLIDADO POR PROCESSO
# ══════════════════════════════════════════════════════════
print('\n4. STF consolidado por processo (amostra 10K)...')
ws = wb.create_sheet('STF Processos')
headers = ['Processo','Classe','Órgão','Relator','Data decisão','Resultado','Tipo decisão','Virtual','Ramo direito','Incidente']
make_header(ws, headers)

rows = query(conn, """
SELECT processo, classe, orgao_julgador, relator_decisao, data_decisao,
  descricao_andamento, tipo_decisao,
  CASE WHEN decisoes_virtual=true THEN 'Virtual' ELSE 'Presencial/Outro' END,
  ramo_direito, incidente
FROM stf_decisoes ORDER BY id LIMIT 10000
""")
write_rows(ws, rows)
auto_width(ws, headers)
print(f'  {len(rows)} rows')

# ══════════════════════════════════════════════════════════
# 5. STF POR AMBIENTE
# ══════════════════════════════════════════════════════════
print('\n5. STF resumo por ambiente...')
ws = wb.create_sheet('STF por Ambiente')
headers = ['Ambiente','Órgão','Total','Unânimes','%Unânimes','Por maioria','Com vencido']
make_header(ws, headers)

rows = query(conn, """
SELECT
  CASE WHEN decisoes_virtual=true THEN 'Virtual' ELSE 'Presencial/Outro' END,
  orgao_julgador, COUNT(*),
  SUM(CASE WHEN observacao_andamento ILIKE '%unanimidade%' THEN 1 ELSE 0 END),
  ROUND(SUM(CASE WHEN observacao_andamento ILIKE '%unanimidade%' THEN 1 ELSE 0 END)*100.0/NULLIF(COUNT(*),0),1),
  SUM(CASE WHEN observacao_andamento ILIKE '%maioria%' THEN 1 ELSE 0 END),
  SUM(CASE WHEN observacao_andamento ILIKE '%vencido%' THEN 1 ELSE 0 END)
FROM stf_decisoes WHERE tipo_decisao='COLEGIADA'
GROUP BY 1, 2 HAVING COUNT(*)>20 ORDER BY 1, 3 DESC
""")
write_rows(ws, rows)
auto_width(ws, headers)
print(f'  {len(rows)} rows')

# ══════════════════════════════════════════════════════════
# 6. STF PARTES CONSOLIDADO
# ══════════════════════════════════════════════════════════
print('\n6. STF partes resumo...')
ws = wb.create_sheet('STF Partes Resumo')
headers = ['Tipo','Papel','Total']
make_header(ws, headers)

rows = query(conn, """
SELECT tipo, papel, COUNT(*) FROM stf_partes
GROUP BY 1, 2 ORDER BY 3 DESC
""")
write_rows(ws, rows)
auto_width(ws, headers)
print(f'  {len(rows)} rows')

# ══════════════════════════════════════════════════════════
# 7. STJ TEMAS COMPLETO
# ══════════════════════════════════════════════════════════
print('\n7. STJ temas completo...')
ws = wb.create_sheet('STJ Temas')
headers = ['Número','Situação','Órgão','Ramo','Relator','Afetação','Julgamento','Trânsito','Tese firmada','Link STF']
make_header(ws, headers)

rows = query(conn, """
SELECT numero, situacao, orgao_julgador, ramo_direito, relator,
  data_afetacao::text, data_julgamento::text, data_transito::text,
  LEFT(tese_firmada, 300), link_stf_rg
FROM stj_temas ORDER BY numero
""")
write_rows(ws, rows)
auto_width(ws, headers)
print(f'  {len(rows)} rows')

# ══════════════════════════════════════════════════════════
# 8. STJ PROCESSOS-SEMENTE
# ══════════════════════════════════════════════════════════
print('\n8. STJ processos-semente...')
ws = wb.create_sheet('STJ Sementes')
headers = ['Tema','Processo','Classe','Número','UF','Tribunal','Relator','RRC','Afetação','Julgamento','Observação']
make_header(ws, headers)

rows = query(conn, """
SELECT tema_numero, processo, classe, numero, uf_origem, tribunal_origem,
  relator, rrc, data_afetacao::text, data_julgamento::text, observacao
FROM stj_processos_semente ORDER BY tema_numero
""")
write_rows(ws, rows)
auto_width(ws, headers)
print(f'  {len(rows)} rows')

# ══════════════════════════════════════════════════════════
# 9. RESUMO EXECUTIVO
# ══════════════════════════════════════════════════════════
print('\n9. Resumo executivo...')
ws = wb.create_sheet('RESUMO EXECUTIVO')
# Move to first position
wb.move_sheet('RESUMO EXECUTIVO', offset=-8)

headers = ['Item', 'Valor', 'Detalhe']
make_header(ws, headers)

# Collect stats
stats = [
    ('', '', ''),
    ('═══ BANCO JUDX ═══', '', ''),
    ('stf_decisoes (raw STF)', query(conn, "SELECT COUNT(*) FROM stf_decisoes")[0][0], 'Corte Aberta 1988-2026'),
    ('judx_case (normalizado)', query(conn, "SELECT COUNT(*) FROM judx_case")[0][0], 'Pipeline completo'),
    ('judx_decision', query(conn, "SELECT COUNT(*) FROM judx_decision")[0][0], 'Pipeline completo'),
    ('stf_partes', query(conn, "SELECT COUNT(*) FROM stf_partes")[0][0], f'{query(conn, "SELECT COUNT(DISTINCT incidente) FROM stf_partes")[0][0]} incidentes'),
    ('stj_temas', query(conn, "SELECT COUNT(*) FROM stj_temas")[0][0], 'Temas repetitivos'),
    ('stj_processos_semente', query(conn, "SELECT COUNT(*) FROM stj_processos_semente")[0][0], 'Com tribunal/relator'),
    ('stj_contramostra', query(conn, "SELECT COUNT(*) FROM stj_contramostra")[0][0], 'CKAN + Datajud'),
    ('judx_court', query(conn, "SELECT COUNT(*) FROM judx_court")[0][0], 'STF + STJ'),
    ('judx_judge', query(conn, "SELECT COUNT(*) FROM judx_judge")[0][0], 'Ministros/juízes'),
    ('judx_subject', query(conn, "SELECT COUNT(*) FROM judx_subject")[0][0], 'Assuntos'),
    ('judx_procedural_class', query(conn, "SELECT COUNT(*) FROM judx_procedural_class")[0][0], 'Classes processuais'),
    ('', '', ''),
    ('═══ ACHADOS CONFIRMADOS ═══', '', ''),
    ('Taxa não-decisão STF', '79%', '169.851 decisões, Corte Aberta'),
    ('Taxa não-decisão STJ', '~90%', 'AREsp 95%, REsp 70.9%'),
    ('Unanimidade virtual', '86,6%', '125.476 de 145.129 colegiadas'),
    ('Processos/ministro/semana (pico)', '~102', 'Semanas de pico 2023-2025'),
    ('Anomalia 2022', '6,1% divergência', 'Metade de qualquer outro ano'),
    ('Marco Aurélio vencido', '8.451 vezes', 'Maior dissidente da história'),
    ('Bloco Mendonça+Nunes', '2.520 juntos (67%)', 'Minoria penal estável'),
    ('Moraes como relator das derrotas do bloco', '54%', '2.022 decisões'),
    ('Temas tributários STJ', '292', '329 dias médio resolução'),
    ('TRF4 como gerador', '388 sementes (15,5%)', 'Maior gerador de temas'),
    ('Circuitos STJ↔STF', '257', 'Via Repercussão Geral'),
    ('', '', ''),
    ('═══ GERADO EM ═══', NOW.strftime('%Y-%m-%d %H:%M'), ''),
]

ri = 2
for item, val, detail in stats:
    c1 = ws.cell(row=ri, column=1, value=item)
    c2 = ws.cell(row=ri, column=2, value=val)
    c3 = ws.cell(row=ri, column=3, value=detail)
    if '═══' in str(item):
        c1.font = SF; c1.fill = SB
        c2.font = SF; c2.fill = SB
        c3.font = SF; c3.fill = SB
    else:
        c1.font = DF; c2.font = NF; c3.font = DF
    ri += 1

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 40

conn.close()

wb.save(OUT)
print(f'\n{"="*50}')
print(f'  INVENTÁRIO SALVO: {OUT}')
print(f'  {len(wb.sheetnames)} abas: {", ".join(wb.sheetnames)}')
print(f'{"="*50}')
