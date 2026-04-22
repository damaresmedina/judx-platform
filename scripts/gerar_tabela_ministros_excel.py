"""Gera versão Excel do stf_175_ministros.csv (tudo já normalizado e corrigido).
Preserva todas as colunas do CSV atual, aplica apenas formatação visual.
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = "C:/Users/medin/Desktop/backup_judx/resultados/stf_175_ministros.csv"
OUT = Path("C:/Users/medin/Desktop/backup_judx/resultados/stf_ministros_trajetoria.xlsx")

with open(SRC, encoding='utf-8') as f:
    rows = list(csv.DictReader(f))
    cols = list(rows[0].keys())

# Ordena por antiguidade crescente (mais antigo primeiro = #1)
rows.sort(key=lambda r: int(r.get('ordem_antiguidade','999')))

# Cabeçalhos legíveis
HEADERS = {
    'ordem_antiguidade': '#',
    'nome_oficial': 'Nome (lista oficial)',
    'nome_completo_bio': 'Nome completo (bio)',
    'nome_curto_bio': 'Nome curto (bio)',
    'nascimento': 'Nascimento',
    'indicacao': 'Indicacao',
    'nomeacao': 'Nomeacao',
    'posse_stf': 'Posse STF',
    'turma_inicial': 'Turma inicial',
    'turma_inicial_from': 'Turma inicial (desde)',
    'trocou_turma': 'Trocou Turma?',
    'turma_1_from': 'Turma 1 (desde)',
    'turma_1_to': 'Turma 1 (ate)',
    'turma_2_from': 'Turma 2 (desde)',
    'turma_2_to': 'Turma 2 (ate)',
    'turma_transicao_data': 'Transicao (data)',
    'turma_transicao_de': 'Transicao (de)',
    'turma_transicao_para': 'Transicao (para)',
    'turma_atual': 'Turma atual',
    'turma_atual_from': 'Turma atual (desde)',
    'foi_pres_turma': 'Foi pres. Turma?',
    'turma_presidida': 'Turma presidida',
    'pres_turma_from': 'Pres. Turma (inicio)',
    'pres_turma_to': 'Pres. Turma (fim)',
    'foi_vice_stf': 'Foi Vice STF?',
    'vice_from': 'Vice STF (inicio)',
    'vice_to': 'Vice STF (fim)',
    'foi_pres_stf': 'Foi Presid. STF?',
    'pres_stf_from': 'Presid. STF (inicio)',
    'pres_stf_to': 'Presid. STF (fim)',
    'aposentadoria': 'Aposentadoria',
    'falecimento': 'Falecimento',
    'casou_bio': 'Bateu com bio?',
}

# Larguras
WIDTHS = {
    'ordem_antiguidade': 5, 'nome_oficial': 52, 'nome_completo_bio': 45, 'nome_curto_bio': 28,
    'nascimento': 14, 'indicacao': 12, 'nomeacao': 12, 'posse_stf': 14,
    'turma_inicial': 11, 'turma_inicial_from': 13, 'trocou_turma': 12,
    'turma_1_from': 13, 'turma_1_to': 13, 'turma_2_from': 13, 'turma_2_to': 13,
    'turma_transicao_data': 13, 'turma_transicao_de': 12, 'turma_transicao_para': 12,
    'turma_atual': 11, 'turma_atual_from': 13,
    'foi_pres_turma': 13, 'turma_presidida': 12, 'pres_turma_from': 13, 'pres_turma_to': 13,
    'foi_vice_stf': 12, 'vice_from': 12, 'vice_to': 12,
    'foi_pres_stf': 13, 'pres_stf_from': 13, 'pres_stf_to': 13,
    'aposentadoria': 13, 'falecimento': 13, 'casou_bio': 12,
}

wb = Workbook()
ws = wb.active
ws.title = "Ministros STF"

# Header
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(bold=True, color='FFFFFF')
for idx, c in enumerate(cols, 1):
    cell = ws.cell(row=1, column=idx, value=HEADERS.get(c, c))
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(idx)].width = WIDTHS.get(c, 15)

# Linhas
for i, r in enumerate(rows, 2):
    for j, c in enumerate(cols, 1):
        v = r.get(c, '')
        cell = ws.cell(row=i, column=j, value=v)
        if c in ('ordem_antiguidade','trocou_turma','foi_pres_turma','foi_vice_stf','foi_pres_stf','casou_bio'):
            cell.alignment = Alignment(horizontal='center')

# Congela cabeçalho + nome
ws.freeze_panes = 'C2'
ws.row_dimensions[1].height = 32

wb.save(OUT)
print(f"[ok] {OUT}")
print(f"linhas: {len(rows)}  |  colunas: {len(cols)}")
