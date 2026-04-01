"""
Extrai todas as normas citadas na CF/88 (planalto.gov.br)
Salva em Excel com 4 abas: Normas, Codigos, ECs, Resumo
"""
import re, os, html as html_mod
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

cf_html = r"C:\Users\medin\Desktop\backup_judx\resultados\cf88_planalto_texto.html"
out_xlsx = r"C:\Users\medin\Desktop\backup_judx\resultados\cf88_base_normativa_2026-03-29.xlsx"

with open(cf_html, "r", encoding="utf-8", errors="replace") as f:
    raw = f.read()

text = re.sub(r"<[^>]+>", " ", raw)
text = html_mod.unescape(text)
text = text.replace("\xa0", " ")
text = re.sub(r"\s+", " ", text)
# Normalizar "n<char especial>" para "no"
text = re.sub(r"n[^\w\s]", "no", text)

print(f"Texto: {len(text)} chars")
print(f"EC: {text.count('Emenda Constitucional no')}")
print(f"Lei: {text.count('Lei no')}")
print(f"LC: {text.count('Lei Complementar no')}")
print(f"DL: {text.count('Decreto-Lei no')}")

patterns = [
    (r"Emenda Constitucional no\s*(\d+)", "EC"),
    (r"Emenda Constitucional de Revis.o no\s*(\d+)", "ECR"),
    (r"Lei Complementar no\s*([\d.]+)", "LC"),
    (r"(?<!Complementar )Lei no\s*([\d.]+)", "Lei"),
    (r"Decreto-Lei no\s*([\d.]+)", "DL"),
    (r"Decreto no\s*([\d.]+)", "Decreto"),
    (r"Medida Provis.ria no\s*([\d.]+)", "MP"),
]

parts = re.split(r"(Art\.\s*\d+[\-A-Z]*)", text)
current_art = "preambulo"
normas = []

for part in parts:
    art_m = re.match(r"Art\.\s*(\d+[\-A-Z]*)", part.strip())
    if art_m:
        current_art = art_m.group(1).replace("-", "")
        continue

    for pat, tipo in patterns:
        for m in re.finditer(pat, part, re.IGNORECASE):
            numero = m.group(1).strip().replace(".", "")
            ctx_after = part[m.end():m.end()+60]
            ano_m = re.search(r"de\s*(\d{4})", ctx_after)
            ano = ano_m.group(1) if ano_m else ""

            ctx = part[max(0, m.start()-80):m.end()+30].lower()
            if "reda" in ctx and "dada" in ctx:
                relacao = "altera_redacao"
            elif "inclu" in ctx:
                relacao = "inclusao"
            elif "revogad" in ctx:
                relacao = "revogacao"
            elif "regulament" in ctx:
                relacao = "regulamenta"
            elif "vide" in ctx:
                relacao = "vide"
            elif "nos termos" in ctx or "na forma" in ctx:
                relacao = "regulamenta"
            else:
                relacao = "referencia"

            normas.append({
                "artigo_cf": f"art-{current_art}",
                "norma_tipo": tipo,
                "norma_numero": numero,
                "norma_ano": ano,
                "tipo_relacao": relacao
            })

seen = set()
unique = []
for n in normas:
    key = (n["artigo_cf"], n["norma_tipo"], n["norma_numero"], n["tipo_relacao"])
    if key not in seen:
        seen.add(key)
        unique.append(n)

print(f"\nTotal: {len(normas)}, unicas: {len(unique)}")
tipos = Counter(n["norma_tipo"] for n in unique)
for t, c in tipos.most_common():
    print(f"  {t}: {c}")

codigos = {
    "10406": "Codigo Civil (2002)",
    "13105": "CPC - Codigo de Processo Civil (2015)",
    "3689": "CPP - Codigo de Processo Penal (1941)",
    "2848": "CP - Codigo Penal (1940)",
    "5452": "CLT (1943)",
    "5172": "CTN - Codigo Tributario Nacional (1966)",
    "8078": "CDC - Codigo de Defesa do Consumidor (1990)",
    "8069": "ECA (1990)",
    "8080": "Lei do SUS (1990)",
    "8112": "Estatuto dos Servidores Federais (1990)",
    "9394": "LDB - Diretrizes e Bases Educacao (1996)",
    "9296": "Interceptacoes Telefonicas (1996)",
    "12527": "Lei de Acesso a Informacao (2011)",
    "13874": "Lei da Liberdade Economica (2019)",
    "8212": "Organizacao da Seguridade Social (1991)",
    "7578": "Pagamento divida em servicos (1986)",
    "9311": "CPMF (1996)",
    "9478": "Lei do Petroleo (1997)",
    "7990": "Compensacao financeira minerais (1989)",
    "12351": "Pre-sal (2010)",
    "14194": "LDO 2022 (2021)",
    "14436": "LDO 2023 (2022)",
    "14601": "Bolsa Familia (2023)",
    "14817": "Piso magistratura (2024)",
}

# Excel
wb = Workbook()
hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(bold=True, color="FFFFFF", size=10)
alt = PatternFill("solid", fgColor="D6E4F0")
thin = Side(style="thin")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)

# ABA 1
ws = wb.active
ws.title = "Normas Citadas"
hdrs = ["Artigo CF", "Tipo", "Numero", "Ano", "Relacao", "Codigo/Lei"]
for i, h in enumerate(hdrs, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd
for i, w in enumerate([14, 8, 12, 8, 16, 45]):
    ws.column_dimensions[get_column_letter(i+1)].width = w
for i, n in enumerate(unique):
    r = i + 2
    cod = codigos.get(n["norma_numero"], "")
    for j, v in enumerate([n["artigo_cf"], n["norma_tipo"], n["norma_numero"], n["norma_ano"], n["tipo_relacao"], cod]):
        c = ws.cell(row=r, column=j+1, value=v)
        c.border = brd
        if i % 2 == 1: c.fill = alt

# ABA 2
ws2 = wb.create_sheet("Codigos e Leis")
for i, h in enumerate(["Codigo/Lei", "Numero", "Tipo", "Artigos CF", "Total"], 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd
ws2.column_dimensions["A"].width = 45
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 8
ws2.column_dimensions["D"].width = 80
ws2.column_dimensions["E"].width = 8
r2 = 2
for num, nome in sorted(codigos.items(), key=lambda x: x[1]):
    arts = sorted(set(n["artigo_cf"] for n in unique if n["norma_numero"] == num))
    if arts:
        tipo = next((n["norma_tipo"] for n in unique if n["norma_numero"] == num), "")
        for j, v in enumerate([nome, num, tipo, ", ".join(arts), len(arts)]):
            ws2.cell(row=r2, column=j+1, value=v).border = brd
        r2 += 1

# ABA 3
ws3 = wb.create_sheet("Emendas Constitucionais")
for i, h in enumerate(["EC", "Ano", "Artigos", "Total", "Relacao principal"], 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = hfont; c.fill = hfill; c.border = brd
ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 8
ws3.column_dimensions["C"].width = 90
ws3.column_dimensions["D"].width = 8
ws3.column_dimensions["E"].width = 16
ecs = {}
for n in unique:
    if n["norma_tipo"] == "EC":
        k = n["norma_numero"]
        if k not in ecs: ecs[k] = {"ano": n["norma_ano"], "arts": set(), "rels": Counter()}
        ecs[k]["arts"].add(n["artigo_cf"])
        ecs[k]["rels"][n["tipo_relacao"]] += 1
r3 = 2
for k in sorted(ecs.keys(), key=lambda x: int(x)):
    ec = ecs[k]
    rel = ec["rels"].most_common(1)[0][0] if ec["rels"] else ""
    for j, v in enumerate([int(k), ec["ano"], ", ".join(sorted(ec["arts"])), len(ec["arts"]), rel]):
        ws3.cell(row=r3, column=j+1, value=v).border = brd
    r3 += 1

# ABA 4
ws4 = wb.create_sheet("Resumo")
lines = [
    ("BASE NORMATIVA DA CRFB/1988", True, 14),
    ("Fonte: planalto.gov.br | 29/03/2026", False, 10),
    ("", False, 10),
    (f"Total normas unicas: {len(unique)}", True, 12),
    ("", False, 10),
    ("POR TIPO:", True, 12),
]
for t, c in tipos.most_common():
    lines.append((f"  {t}: {c}", False, 11))
lines.append(("", False, 10))
lines.append((f"Emendas Constitucionais: {len(ecs)}", True, 12))
lines.append((f"Codigos/Leis principais: {sum(1 for num in codigos if any(n['norma_numero']==num for n in unique))}", True, 12))
for i, (txt, bold, sz) in enumerate(lines):
    ws4.cell(row=i+1, column=1, value=txt).font = Font(bold=bold, size=sz)
ws4.column_dimensions["A"].width = 60

wb.save(out_xlsx)
print(f"\nSalvo: {out_xlsx}")
